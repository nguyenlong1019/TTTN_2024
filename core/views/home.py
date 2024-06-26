from django.shortcuts import render, redirect 
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login

from core.models import * 

from django.db.models import Q, Sum
import json, tempfile
from openpyxl import Workbook 
from django.http import HttpResponse, JsonResponse  
from django.contrib import messages 

from django.core.paginator import Paginator

# Report view
from django.utils import timezone 
from datetime import timedelta 


@login_required(login_url='/login/')
def index(request):
    '''
    Trang chủ:

    Đối với user level 1: Theo dõi hải trình tất cả các tàu
    Đối với user level 2: Theo dõi hải trình của tất cả các tàu thuộc cảng quản lý
    Đối với user level 3: Quản lý thiết bị nhật ký theo dõi
    '''
    if request.user.user_type == '3':
        titles = ["STT", "Serial Number", "Ngày sản xuất", "Version", "Mã tàu", "Trạng thái", "Thao tác"]
        equipments = BangThietBiNhatKyKhaiThac.objects.all().order_by('SerialNumber')
        items_per_page = 7 
        p = Paginator(equipments, items_per_page)
        page = request.GET.get('page')
        items = p.get_page(page)
        current = items.number
        start = max(current - 2, 1)
        end = min(current + 2, items.paginator.num_pages)
        page_range = range(start, end)
        start_number = (current - 1) * items_per_page
        return render(request, 'core/index.html', {
            'titles': titles,
            'items': items,
            'start': start, 
            'end': end, 
            'page_range': page_range,
            'start_number': start_number
        }, status=200)

    if request.user.user_type == '1' or request.user.is_staff:
        ships = []
        # Filter theo tàu nào đã có vị trí, tàu nào chưa có vị trí để tránh lỗi
        ships_with_position = BangTau.objects.prefetch_related('bangvitritau_set')
        ship_counter = 0
        for ship in ships_with_position:
            if ship.bangvitritau_set.exists():
                ships.append(ship)
                ship_counter += 1

        total_ship = len(BangTau.objects.all())
        off_ship = total_ship - ship_counter
        
        return render(request, 'core/index.html', {
            'ships': ships,
            'ship_counter': ship_counter,
            'total_ship': total_ship,
            'off_ship': off_ship,
        }, status=200)
    
    if request.user.user_type == '2':
        ships = []
        ships_with_position = BangTau.objects.filter(CangCaDangKy=request.user.staff.cangca).prefetch_related('bangvitritau_set')
        ship_counter = 0

        for ship in ships_with_position:
            if ship.bangvitritau_set.exists():
                ships.append(ship)
                ship_counter += 1
        
        total_ship = len(BangTau.objects.filter(CangCaDangKy=request.user.staff.cangca))
        off_ship = total_ship - ship_counter

        return render(request, 'core/index.html', {
            'ships': ships,
            'ship_counter': ship_counter,
            'total_ship': total_ship,
            'off_ship': off_ship,
        }, status=200) 

    return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def get_all_location_api(request):
    '''
    Hàm lấy vị trí và thông tin các tàu: chỉ lấy được vị trí các tàu đã có vị trí

    Đối với user level 1: lấy tất cả vị trí và các tàu
    Đối với user level 2: lấy tất cả vị trí và các tàu
    ''' 
    ships = []
    if request.user.user_type == '1' or request.user.is_staff:
        ships_with_position = BangTau.objects.prefetch_related('bangvitritau_set')
    elif request.user.user_type == '2':
        ships_with_position = BangTau.objects.filter(CangCaDangKy=request.user.staff.cangca).prefetch_related('bangvitritau_set')
    else:
        return JsonResponse({
            'message': 'Permission not allowed',
            'status': 403,
            'info': []
        }, status=403) 

    for ship in ships_with_position:
        if ship.bangvitritau_set.exists():
            ships.append(ship)

    info = []
    for i in ships:
        new_loc = i.bangvitritau_set.order_by('-Ngay').first()
        ship_info = {
            'SoDangKy': i.SoDangKy,
            'ViDo': new_loc.ViDo,
            'KinhDo': new_loc.KinhDo,
            'NgayCapNhat': f"{new_loc.Ngay.hour}:{new_loc.Ngay.minute}:{new_loc.Ngay.second} Ngày {new_loc.Ngay.day} tháng {new_loc.Ngay.month} năm {new_loc.Ngay.year}",
        }
        if i.IDChuTau:
            ship_info['ChuTau'] = i.IDChuTau.HoTen
        else:
            ship_info['ChuTau'] = ''
        
        if i.IDThuyenTruong:
            ship_info['ThuyenTruong'] = i.IDThuyenTruong.HoTen
        else:
            ship_info['ThuyenTruong'] = ''
            

        info.append(ship_info)
    return JsonResponse({
        'status': 200,
        'message': 'Cập nhật thông tin vị trí các tàu thành công!',
        'info': info,
    }, status=200)


@login_required(login_url='/login/')
def get_ship_location_api(request, SoDangKy):
    '''
    Hàm lấy vị trí và thông tin của một tàu theo Số đăng ký

    Đối với user level 1: Filter toàn bộ để tìm tàu có Số đăng ký khớp
    Đối với user level 2: Filter theo Số đăng ký và Cảng cá đăng ký của user
    '''
    try:
        if request.user.user_type == '1' or request.user.is_staff:
            ship = BangTau.objects.get(SoDangKy=SoDangKy)
        elif request.user.user_type == '2':
            ship = BangTau.objects.get(SoDangKy=SoDangKy, CangCaDangKy=request.user.staff.cangca)
        else:
            return JsonResponse({
                'message': 'Permission not allowed',
                'status': 403,
                'bundle': {}
            }, status=403)
    except Exception as e:
        return JsonResponse({
            'status': 404,
            'message': 'Ship not Found!',
            'bundle': {}
        }, status=404)

    location = BangViTriTau.objects.filter(IDTau=ship).order_by('-Ngay')[0]
    bundle = {
            'lat': location.ViDo,
            'lng': location.KinhDo,
            'date': f"{location.Ngay.hour}:{location.Ngay.minute}:{location.Ngay.second} Ngày {location.Ngay.day} tháng {location.Ngay.month} năm {location.Ngay.year}",
            'SoDangKy': ship.SoDangKy,
    }
    bundle['shipowner'] = ship.IDChuTau.HoTen if ship.IDChuTau else ''
    bundle['captain'] = ship.IDThuyenTruong.HoTen if ship.IDThuyenTruong else ''
    return JsonResponse({
        'status': 200,
        'message': f'Get info of ship: {ship.SoDangKy} successfully!',
        'bundle': bundle,
    }, status=200)


@login_required(login_url='/login/')
def download_realtime_all_location_view(request):
    if request.user.user_type == '1' or request.user.is_staff:
        ships_with_position = BangTau.objects.prefetch_related('bangvitritau_set')
        ships = []
        for ship in ships_with_position:
            if ship.bangvitritau_set.exists():
                ships.append(ship)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "STT"
        ws['B1'] = "Số đăng ký tàu"
        ws['C1'] = "Chủ tàu"
        ws['D1'] = "Số điện thoại chủ tàu"
        ws['E1'] = "Thuyền trưởng"
        ws['F1'] = "Cảng cá đăng ký"
        ws['G1'] = "Mã thiết bị nhật ký"
        ws['H1'] = "Vĩ độ hiện tại"
        ws['I1'] = "Kinh độ hiện tại"
        ws['J1'] = "Thời gian"

        for i, ship in enumerate(ships, start=2):
            location = ship.bangvitritau_set.order_by('-Ngay').first()
            ws[f"A{i}"] = i - 1
            ws[f"B{i}"] = ship.SoDangKy
            ws[f"C{i}"] = ship.IDChuTau.HoTen if ship.IDChuTau else ''
            ws[f"D{i}"] = ship.IDChuTau.DienThoai if ship.IDChuTau else ''
            ws[f"E{i}"] = ship.IDThuyenTruong.HoTen if ship.IDThuyenTruong else ''
            ws[f"F{i}"] = ship.CangCaDangKy.Ten
            ws[f"G{i}"] = ship.IDDevice.SerialNumber if ship.IDDevice else ''
            ws[f"H{i}"] = location.ViDo
            ws[f"I{i}"] = location.KinhDo
            ws[f"J{i}"] = f"{location.Ngay.hour}:{location.Ngay.minute}:{location.Ngay.second} {location.Ngay.day}-{location.Ngay.month}-{location.Ngay.year}"
            

        # Trước khi lưu workbook, tính toán chiều rộng tối ưu cho từng cột
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            # Cài đặt chiều rộng của cột
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2  # Thêm một giới hạn để tránh chữ bị cắt

        _, filepath = tempfile.mkstemp(suffix='.xlsx')
        wb.save(filepath)

        with open(filepath, 'rb') as f:
            excel_data = f.read()

        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="realtime-location.xlsx"'
        return response

    else:
        return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def change_password_view(request):
    if request.method == 'POST':
        password = request.POST.get('password')
        user = request.user 
        user.set_password(password)
        user.save()
        
        login(request, user)

        messages.info(request, "Cập nhật mật khẩu thành công!")
        return redirect('index')
    
    return render(request, 'change-password.html', {})


@login_required(login_url='/login/')
def marine_diary_view(request):
    """
    Lịch sử hải trình
    """
    query = request.GET.get('q')
    start_date = request.GET.get('start-date')
    end_date = request.GET.get('end-date')
    if request.user.user_type == '1' or request.user.is_staff:
        try:
            ships = BangTau.objects.filter(Q(SoDangKy__icontains=query))
            if ships.exists():
                ship = BangTau.objects.filter(Q(SoDangKy__icontains=query)).first()
                if not ship.bangvitritau_set.exists():
                    messages.info(request, f"Thông tin tàu {ship.SoDangKy} chưa được cập nhật vị trí")
                    return redirect('index')
            else:
                messages.info(request, f"Không tìm thấy tàu với query = '{query}'") 
                return redirect('index')        
        except ValueError:
            messages.info(request, f"Số đăng ký tàu không được phép trống hoặc None!!")
            return redirect('index')

        return render(request, 'core/index-search.html', {
            'ship': ship,
            'start_date': start_date,
            'end_date': end_date
        }, status=200)
    elif request.user.user_type == '2':
        try:
            ships = BangTau.objects.filter(Q(SoDangKy__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca))
            if ships.exists():
                ship = BangTau.objects.filter(Q(SoDangKy__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca)).first()
                if not ship.bangvitritau_set.exists():
                    messages.info(request, f"Thông tin tàu {ship.SoDangKy} chưa được cập nhật vị trí")
                    return redirect('index')
            else:
                messages.info(request, f"Không tìm thấy tàu với query = '{query}' tại cảng cá {request.user.staff.cangca.Ten}") 
                return redirect('index')        
        except ValueError:
            messages.info(request, f"Số đăng ký tàu không được phép trống hoặc None!!")
            return redirect('index')

        return render(request, 'core/index-search.html', {
            'ship': ship,
            'start_date': start_date,
            'end_date': end_date
        }, status=200) 
    else:
        return render(request, '403.html', {}, status=403)


# lấy lịch sử vị trí tàu 
@login_required(login_url='/login/')
def get_history_ship_location_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pk = data.get('pk')
            start_date = data.get('startDate');
            end_date = data.get('endDate')

            try:
                ship = BangTau.objects.get(pk=pk)
            except BangTau.DoesNotExist:
                messages.info(request, f"Không tìm thấy thông tin tàu với query = '{pk}'")
                return redirect('index')

            location = BangViTriTau.objects.filter(
                Q(Ngay__gte=start_date) & Q(Ngay__lte=end_date),
                IDTau = ship
            ).order_by('Ngay').values('ViDo', 'KinhDo', 'Ngay')
            # print(location)
            if location.exists():
                for loc in location:
                    loc['Ngay'] = loc['Ngay'].strftime('%Y-%m-%d %H:%M:%S')
                # print(location)

                location_data = list(location)

                info = {
                    'SoDangKy': ship.SoDangKy,
                    'ChuTau': ship.IDChuTau.HoTen if ship.IDChuTau else '',
                    'ThuyenTruong': ship.IDThuyenTruong.HoTen if ship.IDThuyenTruong else '',
                    'ViDo': location_data[-1]['ViDo'],
                    'KinhDo': location_data[-1]['KinhDo'],
                    'Ngay': location_data[-1]['Ngay']
                }


                return JsonResponse({
                    'message': f'Nhật ký hải trình tàu {ship.SoDangKy}',
                    'status': 200,
                    'location': location_data,
                    'info': info,
                }, status=200)

            else:
                return JsonResponse({
                    'message': f'Không tìm thấy thông tin vị trí!',
                    'status': 404,
                }, status=404) 
        
        except json.JSONDecodeError:
            return JsonResponse({
                'message': 'Dữ liệu không hợp lệ',
                'status': 400
            }, status=400)
    else:
        return JsonResponse({
            'message': 'Method not allowed',
            'status': 405
        }, status=405)

