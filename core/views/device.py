from django.shortcuts import render, redirect 
from django.contrib.auth.decorators import login_required 

from core.models import * 

from django.db.models import Q 
from openpyxl import Workbook 
import tempfile 
from django.http import HttpResponse, JsonResponse 
from django.contrib import messages

# Import Pagination Stuff 
from django.core.paginator import Paginator

 
@login_required(login_url='/login/')
def device_view(request):
    '''Quản lý tàu cá
    
    Sort theo số đăng ký
    Phân trang với 7 tàu trên 1 trang
    '''
    titles = ["STT","IMO","Số hiệu tàu","Chủ tàu","Loại tàu","Nơi đăng ký","Thuyền trưởng","Cảng cá ĐK","TBNKKT","Thao tác"]
    
    if request.user.user_type == '1' or request.user.is_staff:
        # items = BangTau.objects.all().order_by('SoDangKy')

        # Set up pagination: p for pagination: 7 items on page
        p = Paginator(BangTau.objects.all().order_by('SoDangKy'), 7)
    elif request.user.user_type == '2':
        # items = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca)).order_by('SoDangKy')
        p = Paginator(BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca)).order_by('SoDangKy'), 7)
    else:
        return render(request, '403.html', {}, status=403)
    
    # p = Paginator(items, 8)
    page = request.GET.get('page')
    ships = p.get_page(page)
    # nums = [i+1 for i in range(ships.paginator.num_pages)]
    # print(nums)
    current = ships.number # current page
    start = max(current - 2, 1)
    end = min(current + 2, ships.paginator.num_pages)

    page_range = range(start, end)

    items_per_page = 7
    start_number = (current - 1) * items_per_page

    return render(request, 'core/device.html', {
        'titles': titles, 
        'ships': ships,
        'start': start, 
        'end': end, 
        'page_range': page_range,
        'start_number': start_number
    }, status=200)


def check_id_valid(id):
    '''Kiểm tra số đăng ký tàu đã có hay chưa?'''
    ship = BangTau.objects.filter(SoDangKy=id)
    if ship is not None:
        return False
    return True 


@login_required(login_url='/login/')
def add_new_device_view(request):
    '''Thêm tàu cá mới'''
    if request.user.user_type == '1' or request.user.is_staff or request.user.user_type == '2':
        if request.method == 'POST':
            TenTau = request.POST.get('tenTau')
            try:
                ChuTau = BangChuTau.objects.get(pk=request.POST['chuTau'])
            except Exception as e:
                messages.info(request, "Chủ tàu không tồn tại!!!")
                return redirect('add-device')

            try:
                LoaiTau = BangMaLoaiTau.objects.get(pk=request.POST['loaiTau'])
            except Exception as e:
                messages.info(request, "Loại tàu không tồn tại!!!")
                return redirect('add-device')

            SoDangKy = request.POST.get('soDangKy')
            # Xử lý số đăng ký trùng hay không
            if not check_id_valid(SoDangKy):
                messages.info(request, "Số đăng ký tàu đã tồn tại!!!")
                return redirect('add-device')

            HoHieu = request.POST.get('hoHieu')
            CoHieu = request.POST.get('coHieu')
            IMO = request.POST.get('IMO')
            
            try:
                NoiDangKy = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['noiDangKy'])
            except Exception as e:
                messages.info(request, "Nơi đăng ký không tồn tại!!!")
                return redirect('add-device')

            try:
                CangCaDangKy = BangCangCa.objects.get(pk=request.POST['cangCaDangKy'])
            except Exception as e:
                messages.info(request, "Thông tin cảng cá không tồn tại!!!")
                return redirect('add-device')
            
            try:
                CangCaPhu = BangCangCa.objects.get(pk=request.POST['cangCaPhu'])
            except:
                messages.info(request, "Thông tin cảng cá phụ tồn tại!!!")
                return redirect('add-device')

            try:
                NgheChinh = BangNganhNgheKhaiThac.objects.get(pk=request.POST['ngheChinh'])
            except Exception as e:
                messages.info(request, "Ngành nghề không tồn tại!!!")
                return redirect('add-device')
            
            NghePhu1 = BangNganhNgheKhaiThac.objects.get(pk=request.POST['nghePhu1'])
            NgayDangKy = request.POST.get('ngayDangKy')
            NgayHetHanDangKy = request.POST.get('ngayHetHanDangKy')
            NgaySanXuatTau = request.POST.get('ngaySanXuatTau')
            NgayHetHanSuDung = request.POST.get('ngayHetHanSuDung')
            
            try:
                MaThietBi = BangThietBiNhatKyKhaiThac.objects.get(pk=request.POST['maThietBi'])
            except Exception as e:
                messages.info(request, "Mã thiết bị không tồn tại!!!")
                return redirect('add-device')
            
            try:
                ThuyenTruong = BangThuyenTruong.objects.get(pk=request.POST['thuyenTruong'])
            except Exception as e:
                messages.info(request, "Thông tin thuyền trưởng không tồn tại!!!")
                return redirect('add-device')
            
            try:
                Tinh = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['tinhThanhPho'])
            except Exception as e:
                messages.info(request, "Thông tin tỉnh không tồn tại!!!")
                return redirect('add-device')
            
            SoLuongThuyenVien = request.POST.get('soLuongThuyenVien')
            TongTaiTrong = request.POST.get('tongTaiTrong')
            ChieuDaiLonNhat = request.POST.get('chieuDaiLonNhat')
            ChieuRongLonNhat = request.POST.get('chieuRongLonNhat')
            CongSuatMay = request.POST.get('congSuatMay')
            MonNuoc = request.POST.get('monNuoc')
            DungTichHamCa = request.POST.get('dungTichHamCa')
            VanTocDanhBat = request.POST.get('vanTocDanhBat')
            VanTocHanhTrinh = request.POST.get('vanTocHanhTrinh')
            ThongSoNguCu = request.POST.get('thongSoNguCu')

            new_ship = BangTau.objects.create(
                SoDangKy=SoDangKy,
                TenTau=TenTau,
                HoHieu=HoHieu,
                CoHieu=CoHieu,
                IMO=IMO,
                NoiDangKy=NoiDangKy,
                CangCaDangKy=CangCaDangKy,
                NgheChinh=NgheChinh,
                NghePhu1=NghePhu1,
                NgayDangKy=NgayDangKy,
                NgayHetHanDangKy=NgayHetHanDangKy,
                TongTaiTrong=TongTaiTrong,
                ChieuDaiLonNhat=ChieuDaiLonNhat,
                ChieuRongLonNhat=ChieuRongLonNhat,
                CongSuatMay=CongSuatMay,
                MonNuoc=MonNuoc,
                SoThuyenVien=SoLuongThuyenVien,
                NgaySanXuat=NgaySanXuatTau,
                NgayHetHan=NgayHetHanSuDung,
                LoaiTau=LoaiTau,
                DungTichHamCa=DungTichHamCa,
                VanTocDanhBat=VanTocDanhBat,
                VanTocHanhTrinh=VanTocHanhTrinh,
                IDDevice=MaThietBi,
                IDChuTau=ChuTau,
                IDThuyenTruong=ThuyenTruong,
                IDTinh=Tinh,
                ThongSoNguCu=ThongSoNguCu
            )

            MaThietBi.is_active = True
            MaThietBi.save()
            ThuyenTruong.hasShip = True
            ThuyenTruong.save()

            messages.success(request, f"Thêm thông tin tàu {SoDangKy} thành công!!")
            return redirect('device-view')


        # GET METHOD

        city_list = BangDonViHanhChinhCapTinh.objects.all().order_by('TenTiengViet')
        shipowners = BangChuTau.objects.all().order_by('HoTen')
        ship_type_list = BangMaLoaiTau.objects.all().order_by('IDLoaiTau')
        gate_list = BangCangCa.objects.all().order_by('Ten')
        job_list = BangNganhNgheKhaiThac.objects.all().order_by('Ten')
        device_list = BangThietBiNhatKyKhaiThac.objects.all().order_by('SerialNumber').exclude(is_active=True) # danh sách thiết bị chưa kích hoạt
        captain_list = BangThuyenTruong.objects.all().order_by('HoTen').exclude(hasShip=True) # danh sách các thuyền trưởng chưa có tàu nào
        today = date.today()

        return render(request, 'core/add-new-device.html', {
            'shipowners': shipowners,
            'ship_type_list': ship_type_list,
            'city_list': city_list,
            'gate_list': gate_list,
            'job_list': job_list,
            'device_list': device_list,
            'captain_list': captain_list,
            'today': today,
        }, status=200)
    
    else:
        return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def edit_device_view(request, pk):
    if request.user.user_type == '2' or request.user.user_type == '1' or request.user.is_staff:
        if request.method == 'POST':
            TenTau = request.POST.get('tenTau')
            try:
                ChuTau = BangChuTau.objects.get(pk=request.POST['chuTau'])
            except Exception as e:
                messages.info(request, "Thông tin chủ tàu không tồn tại!!!")
                return redirect('edit-device', pk)

            try:
                LoaiTau = BangMaLoaiTau.objects.get(pk=request.POST['loaiTau'])
            except Exception as e:
                messages.info(request, "Thông tin mã loại tàu không tồn tại!!!")
                return redirect('edit-device', pk)

            SoDangKy = request.POST.get('soDangKy') 
            # xử lý số đăng ký trùng tại phần kiểm tra xem số đăng ký là số đăng ký cũ hay mới
            

            HoHieu = request.POST.get('hoHieu')
            CoHieu = request.POST.get('coHieu')
            IMO = request.POST.get('IMO')
            
            try:
                NoiDangKy = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['noiDangKy'])
            except Exception as e:
                messages.info(request, "Thông tin nơi đăng ký không tồn tại!!!")
                return redirect('edit-device', pk)
            
            try:
                CangCaDangKy = BangCangCa.objects.get(pk=request.POST['cangCaDangKy'])
            except Exception as e:
                messages.info(request, "Thông tin cảng cá đăng ký không tồn tại!!!")
                return redirect('edit-device', pk)
            
            try:
                CangCaPhu = BangCangCa.objects.get(pk=request.POST['cangCaPhu'])
            except:
                messages.info(request, "Thông tin cảng cá phụ tồn tại!!!")
                return redirect('edit-device', pk)

            try:
                NgheChinh = BangNganhNgheKhaiThac.objects.get(pk=request.POST['ngheChinh'])
            except Exception as e:
                messages.info(request, "Thông tin nghề chính không tồn tại!!!")
                return redirect('edit-device', pk)
            
            NghePhu1 = BangNganhNgheKhaiThac.objects.get(pk=request.POST['nghePhu1'])
            NgayDangKy = request.POST.get('ngayDangKy')
            NgayHetHanDangKy = request.POST.get('ngayHetHanDangKy')
            NgaySanXuatTau = request.POST.get('ngaySanXuatTau')
            NgayHetHanSuDung = request.POST.get('ngayHetHanSuDung')
            
            try:
                MaThietBi = BangThietBiNhatKyKhaiThac.objects.get(pk=request.POST['maThietBi'])
            except Exception as e:
                messages.info(request, "Thông tin mã thiết bị không tồn tại!!!")
                return redirect('edit-device', pk)
            
            try:
                ThuyenTruong = BangThuyenTruong.objects.get(pk=request.POST['thuyenTruong'])
            except Exception as e:
                messages.info(request, "Thông tin thuyền trưởng không tồn tại!!!")
                return redirect('edit-device', pk)
            
            try:
                Tinh = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['tinhThanhPho'])
            except Exception as e:
                messages.info(request, "Thông tin tỉnh không tồn tại!!!")
                return redirect('edit-device', pk)
            
            SoLuongThuyenVien = request.POST.get('soLuongThuyenVien')
            TongTaiTrong = request.POST.get('tongTaiTrong')
            ChieuDaiLonNhat = request.POST.get('chieuDaiLonNhat')
            ChieuRongLonNhat = request.POST.get('chieuRongLonNhat')
            CongSuatMay = request.POST.get('congSuatMay')
            MonNuoc = request.POST.get('monNuoc')
            DungTichHamCa = request.POST.get('dungTichHamCa')
            VanTocDanhBat = request.POST.get('vanTocDanhBat')
            VanTocHanhTrinh = request.POST.get('vanTocHanhTrinh')
            ThongSoNguCu = request.POST.get('thongSoNguCu')

            try:
                ship = BangTau.objects.get(pk=pk)
            except Exception as e:
                messages.info(request, "Thông tin tàu không tồn tại!!!")
                return redirect('edit-device', pk)

            if ship.SoDangKy != SoDangKy:
                # Xử lý số đăng ký trùng hay không
                if not check_id_valid(SoDangKy):
                    messages.info(request, "Số đăng ký tàu đã tồn tại!!!")
                    return redirect('edit-device', pk)
                else:
                    ship.SoDangKy=SoDangKy

            tbnkkt = ship.IDDevice
            tbnkkt.is_active = False
            tbnkkt.save()

            captain = ship.IDThuyenTruong
            captain.hasShip = False
            captain.save()

            ship.TenTau=TenTau
            ship.HoHieu=HoHieu
            ship.CoHieu=CoHieu
            ship.IMO=IMO
            ship.NoiDangKy=NoiDangKy
            ship.CangCaDangKy = CangCaDangKy
            ship.NgheChinh=NgheChinh
            ship.NghePhu1=NghePhu1
            ship.NgayDangKy=NgayDangKy
            ship.NgayHetHanDangKy=NgayHetHanDangKy
            ship.TongTaiTrong=TongTaiTrong
            ship.ChieuDaiLonNhat=ChieuDaiLonNhat
            ship.ChieuRongLonNhat=ChieuRongLonNhat
            ship.CongSuatMay=CongSuatMay
            ship.MonNuoc=MonNuoc
            ship.SoThuyenVien=SoLuongThuyenVien
            ship.NgaySanXuat=NgaySanXuatTau
            ship.NgayHetHan=NgayHetHanSuDung
            ship.LoaiTau=LoaiTau
            ship.DungTichHamCa=DungTichHamCa
            ship.VanTocDanhBat=VanTocDanhBat
            ship.VanTocHanhTrinh=VanTocHanhTrinh
            ship.IDDevice=MaThietBi
            ship.IDChuTau=ChuTau
            ship.IDThuyenTruong=ThuyenTruong
            ship.IDTinh=Tinh
            ship.ThongSoNguCu=ThongSoNguCu

            MaThietBi.is_active = True
            MaThietBi.save()
            ThuyenTruong.hasShip = True
            ThuyenTruong.save()

            ship.save()    
            messages.success(request, f"Cập nhật thông tin tàu {ship.SoDangKy} thành công!!")
            return redirect('device-view')


        try:
            item = BangTau.objects.get(pk=pk)
        except Exception as e:
            messages.info(request, f'Không tìm thấy thông tin tàu hợp lệ!')
            redirect('device-view')
        
        shipowners = BangChuTau.objects.all().order_by('HoTen')
        ship_type_list = BangMaLoaiTau.objects.all().order_by('IDLoaiTau')
        city_list = BangDonViHanhChinhCapTinh.objects.all().order_by('TenTiengViet')
        gate_list = BangCangCa.objects.all().order_by('Ten')
        job_list = BangNganhNgheKhaiThac.objects.all().order_by('Ten')
        device_list = BangThietBiNhatKyKhaiThac.objects.all().order_by('SerialNumber').exclude(is_active=True)
        captain_list = BangThuyenTruong.objects.all().order_by('HoTen').exclude(hasShip=True)

        return render(request, 'core/edit-device.html', {
            'item': item,
            'shipowners': shipowners,
            'ship_type_list': ship_type_list,
            'city_list': city_list,
            'gate_list': gate_list,
            'job_list': job_list,
            'device_list': device_list,
            'captain_list': captain_list,
        }, status=200)
    else:
        return render(request, '403.html', {}, status=403) 


@login_required(login_url='/login/')
def delete_device_view(request, pk):
    if request.user.user_type == '1' or request.user.is_staff or request.user.user_type == '2':
        try:
            ship = BangTau.objects.get(pk=pk)
            soDangKy = ship.SoDangKy

            tbnkkt = ship.IDDevice
            tbnkkt.is_active = False
            tbnkkt.save()

            captain = ship.IDThuyenTruong
            captain.hasShip = False
            captain.save()
            
            ship.delete()
            messages.success(request, f"Xóa tàu {soDangKy} thành công!!")
            return redirect('device-view')
        except Exception as e:
            messages.error(request, f"Tàu mới pk = {pk} không tồn tại!!")
            return redirect('device-view')
    else:
        return render(request, '403.html', {}, status=403) 


@login_required(login_url='/login/')
def search_device_view(request):
    query = request.GET.get('q')
    query_type = request.GET.get('query-type')
    titles = ["STT","IMO","Số hiệu tàu","Chủ tàu","Loại tàu","Nơi đăng ký","Thuyền trưởng","Cảng cá ĐK","TBNKKT","Thao tác"]
    
    if request.user.user_type == '1' or request.user.is_staff:
        if query_type == '1':
            items = BangTau.objects.filter(Q(IMO__icontains=query)).order_by('IMO')
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy IMO hợp lệ!')
        elif query_type == '2':
            items = BangTau.objects.filter(Q(SoDangKy__icontains=query)).order_by('SoDangKy') 
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy số đăng ký tàu hợp lệ!')
        elif query_type == '3':
            users = BangChuTau.objects.filter(Q(HoTen__icontains=query)).order_by('HoTen')
            items = []
            for user in users:
                item = user.bangtau_chutau.all()
                items.extend(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy chủ tàu hợp lệ!')
        elif query_type == '4':
            users = BangThuyenTruong.objects.filter(Q(HoTen__icontains=query)).order_by('HoTen').exclude(hasShip=False)
            items = []
            for user in users:
                print(user)
                item = user.bangtau
                items.append(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thuyền trưởng hợp lệ!')
        elif query_type == '5':
            devices = BangThietBiNhatKyKhaiThac.objects.filter(Q(SerialNumber__icontains=query)).order_by('SerialNumber').exclude(is_active=False)
            items = []
            for d in devices:
                item = d.bangtau
                items.append(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thiết bị hợp lệ!')
        elif query_type == '6':
            items = BangTau.objects.filter(Q(IMO__icontains=query) | Q(SoDangKy__icontains=query)).order_by('SoDangKy')  
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin IMO hoặc Số đăng ký hợp lệ!')
        else:
            items = BangTau.objects.all().order_by('SoDangKy') 

    elif request.user.user_type == '2':
        if query_type == '1':
            items = BangTau.objects.filter(Q(IMO__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca)).order_by('IMO')
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy IMO hợp lệ!')
        elif query_type == '2':
            items = BangTau.objects.filter(Q(SoDangKy__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca)).order_by('SoDangKy') 
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy số đăng ký tàu hợp lệ!')
        elif query_type == '3':
            users = BangChuTau.objects.filter(Q(HoTen__icontains=query)).order_by('HoTen')
            items = []
            for user in users:
                item = user.bangtau_chutau.all()
                items.extend(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy chủ tàu hợp lệ!')
        elif query_type == '4':
            users = BangThuyenTruong.objects.filter(Q(HoTen__icontains=query)).order_by('HoTen').exclude(hasShip=False)
            items = []
            for user in users:
                print(user)
                item = user.bangtau
                if item.CangCaDangKy.ID == request.user.staff.cangca.ID:
                    items.append(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thuyền trưởng hợp lệ!')
        elif query_type == '5':
            devices = BangThietBiNhatKyKhaiThac.objects.filter(Q(SerialNumber__icontains=query)).order_by('SerialNumber').exclude(is_active=False)
            items = []
            for d in devices:
                item = d.bangtau
                if item.CangCaDangKy.ID == request.user.staff.cangca.ID:
                    items.append(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thiết bị hợp lệ!')
        elif query_type == '6':
            items = BangTau.objects.filter((Q(IMO__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca)) | (Q(SoDangKy__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca))).order_by('SoDangKy')  
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin IMO hoặc Số đăng ký hợp lệ!')
        else:
            items = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca)).order_by('HoTen')
    
    else:
        return render(request, '403.html', {}, status=403)
    
    p = Paginator(items, 7)
    page = request.GET.get('page')
    ships = p.get_page(page)
    current = ships.number # current page
    start = max(current - 2, 1)
    end = min(current + 2, ships.paginator.num_pages)

    page_range = range(start, end)

    items_per_page = 7
    start_number = (current - 1) * items_per_page

    return render(request, 'core/device.html', {
        'titles': titles,
        'ships': ships,
        'start': start, 
        'end': end, 
        'page_range': page_range,
        'start_number': start_number 
    }, status=200)


@login_required(login_url='/login/')
def download_device_data(request, number):
    number = int(number)
    wb = Workbook()
    ws = wb.active # worksheet
    if number == '1':
        qty = 50
    elif number == '2':
        qty = 100
    elif number == '3':
        qty = 200
    elif number == '4':
        qty = 500
    else:
        qty = None 

    if request.user.user_type == '1' or request.user.is_staff:
        if qty is not None:
            ship_list = BangTau.objects.all()[:qty]
        else:
            ship_list = BangTau.objects.all()
    elif request.user.user_type == '2':
        if qty is not None:
            ship_list = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca))[:qty]
        else:
            ship_list = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca))
    else:
        return render(request, '403.html', {}, status=403) 

    ws["A1"] = "STT"
    ws["B1"] = "IMO"
    ws["C1"] = "Số hiệu tàu"
    ws["D1"] = "Chủ tàu"
    ws["E1"] = "Số điện thoại chủ tàu"
    ws["F1"] = "CMND/CCCD" 
    ws["G1"] = "Thuyền trưởng"
    ws["H1"] = "Cảng cá ĐK"
    ws["I1"] = "TBNKKT" 
    ws["J1"] = "Loại tàu"
    ws["K1"] = "Nơi đăng ký"

    for i, ship in enumerate(ship_list, start=2):
        if ship is not None:
            ws[f"A{i}"] = i - 1
            ws[f"B{i}"] = ship.IMO
            if request.user.user_type == '2':
                ws[f"C{i}"] = "********"
            else:
                ws[f"C{i}"] = ship.SoDangKy
            ws[f"D{i}"] = ship.IDChuTau.HoTen
            ws[f"E{i}"] = ship.IDChuTau.DienThoai
            ws[f"F{i}"] = ship.IDChuTau.CMND_CCCD
            ws[f"G{i}"] = ship.IDThuyenTruong.HoTen
            ws[f"H{i}"] = ship.CangCaDangKy.Ten
            ws[f"I{i}"] = ship.IDDevice.SerialNumber
            ws[f"J{i}"] = ship.LoaiTau.IDLoaiTau
            ws[f"K{i}"] = ship.NoiDangKy.TenTiengViet
        else:
            ws[f"A{i}"] = ''
            ws[f"B{i}"] = ''
            ws[f"C{i}"] = ''
            ws[f"D{i}"] = ''
            ws[f"E{i}"] = ''
            ws[f"F{i}"] = ''
            ws[f"G{i}"] = ''
            ws[f"H{i}"] = ''
            ws[f"I{i}"] = ''
            ws[f"J{i}"] = ''
            ws[f"K{i}"] = ''

    # Trước khi lưu workbook, tính toán chiều rộng tối ưu cho từng cột
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        # Cài đặt chiều rộng của cột
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2  # Thêm một giới hạn để tránh chữ bị cắt

    _, filepath = tempfile.mkstemp(suffix='.xlsx')
    wb.save(filepath)

    with open(filepath, 'rb') as f:
        excel_data = f.read()
    
    response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ships_data.xlsx"'
    return response
    