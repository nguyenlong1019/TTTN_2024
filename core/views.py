from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate 
from django.http import JsonResponse, HttpResponse
import json 
import os 
from dotenv import load_dotenv # pip install python-dotenv
from django.views.decorators.csrf import csrf_exempt
import json, random, operator
from datetime import datetime, date 
from .models import * 
from django.contrib.auth.models import User 

# Get CustomUser 
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password

from django.contrib import messages
from django.db.models import Q, Max, Sum
from django.http import HttpResponse 
from openpyxl import Workbook 
import tempfile

# Reportlab for PDF
from reportlab.platypus import Paragraph, Frame, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas 
from reportlab.lib.units import cm 
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle 
from reportlab.lib.pagesizes import letter, landscape 
from reportlab.pdfbase import pdfmetrics 
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors


# Report view
from django.db.models import Sum 
from datetime import datetime, timedelta
from django.utils import timezone
from django.utils.timezone import make_aware


# Handle 404, 500, 400, 403
def handler404(request, *args, **argv):
    return render(request, '404.html', {}, status=404)  


def handler500(request, *args, **argv):
    return render(request, '500.html', {}, status=500)


def handler400(request, *args, **argv):
    return render(request, '400.html', {}, status=400)


def handler403(request, *args, **argv):
    return render(request, '403.html', {}, status=403)


# Login
def login_view(request):
    if request.user.is_authenticated:
        return redirect('index')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('index')
                # print(user.user_type, type(user.user_type))
                if user.user_type == '3':
                    return redirect('provider-home') 
                else:
                    return redirect('index')

            else:
                return render(request, 'core/login.html', {'message': 'Password không đúng!!'}, status=401)
        except Exception as e:
            return render(request, 'core/login.html', {'message': f'User với {username} không tồn tại!'}, status=500)
    return render(request, 'core/login.html', status=200)


# Logout 
@login_required(login_url='/login/')
def logout_view(request):
    logout(request)
    return redirect('login')


# @login_required(login_url='/login/')
# def provider_home_view(request):
#     if request.user.user_type == '3':
#         return HttpResponse("User 3")
#     else:
#         return HttpResponse("Không có quyền truy cập vào trang này!")


# Quản lý hải trình
@login_required(login_url='/login/')
def index(request):
    '''Quản lý hải trình'''
    if request.user.user_type == '3':
        titles = ["STT", "Serial Number", "Ngày sản xuất", "Version", "Mã tàu", "Trạng thái", "Thao tác"]
        equipments = BangThietBiNhatKyKhaiThac.objects.all().order_by('-ID')
        return render(request, 'core/index.html', {
            'titles': titles,
            'items': equipments
        }, status=200)
    if request.user.user_type == '1' or request.user.is_staff:

        load_dotenv()
        API_KEY = os.getenv("GOOGLE_MAP_API_KEY")

        # Filter theo tàu nào đã có vị trí, tàu nào chưa có vị trí để tránh lỗi
        # ships = BangTau.objects.all()
        ships = []
        ships_with_position = BangTau.objects.prefetch_related('bangvitritau_set')
        ship_counter = 0
        for ship in ships_with_position:
            if ship.bangvitritau_set.exists():
                ships.append(ship)
                ship_counter += 1
        

        return render(request, 'core/index.html', {
            'api_key': API_KEY,
            'ships': ships,
            'ship_counter': ship_counter,
        }, status=200)
    
    if request.user.user_type == '2':
        ships = []
        # staff = request.user.staff
        ships_with_position = BangTau.objects.filter(CangCaDangKy=request.user.staff.cangca).prefetch_related('bangvitritau_set')
        ship_counter = 0
        for ship in ships_with_position:
            if ship.bangvitritau_set.exists():
                ships.append(ship)
                ship_counter += 1

        return render(request, 'core/index.html', {
            'ships': ships,
            'ship_counter': ship_counter
        }, status=200) 

    return render(request, '403.html', {}, status=403)


# Lịch sử hải trình
@login_required(login_url='/login/')
def marine_diary_view(request):
    query = request.GET.get('q')
    start_date = request.GET.get('start-date')
    end_date = request.GET.get('end-date')
    if request.user.user_type == '1' or request.user.is_staff:

        # hiểu, vì filter nó sẽ không gây ra DoesNotExist
        try:
            ships = BangTau.objects.filter(Q(SoDangKy__icontains=query))
            if ships.exists():
                ship = BangTau.objects.filter(Q(SoDangKy__icontains=query)).first()
                if not ship.bangvitritau_set.exists():
                    messages.info(request, f"Thông tin tàu {ship.SoDangKy} chưa được cập nhật vị trí")
                    return redirect('index')
            else:
                messages.info(request, f"Không tìm thấy tàu với query = '{query}'") 
                return redirect('index')        
        except ValueError:
            messages.info(request, f"Số đăng ký tàu không được phép trống hoặc None!!")
            return redirect('index')

        return render(request, 'core/index-search.html', {
            'ship': ship,
            'start_date': start_date,
            'end_date': end_date
        }, status=200)
    elif request.user.user_type == '2':
        try:
            ships = BangTau.objects.filter(Q(SoDangKy__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca))
            if ships.exists():
                ship = BangTau.objects.filter(Q(SoDangKy__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca)).first()
                if not ship.bangvitritau_set.exists():
                    messages.info(request, f"Thông tin tàu {ship.SoDangKy} chưa được cập nhật vị trí")
                    return redirect('index')
            else:
                messages.info(request, f"Không tìm thấy tàu với query = '{query}' tại cảng cá {request.user.staff.cangca.Ten}") 
                return redirect('index')        
        except ValueError:
            messages.info(request, f"Số đăng ký tàu không được phép trống hoặc None!!")
            return redirect('index')

        return render(request, 'core/index-search.html', {
            'ship': ship,
            'start_date': start_date,
            'end_date': end_date
        }, status=200) 
    else:
        return render(request, '403.html', {}, status=403)




# lấy lịch sử vị trí tàu 
@login_required(login_url='/login/')
def get_ship_location_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pk = data.get('pk')
            start_date = data.get('startDate');
            end_date = data.get('endDate')

            try:
                ship = BangTau.objects.get(pk=pk)
            except BangTau.DoesNotExist:
                messages.info(request, f"Không tìm thấy thông tin tàu với query = '{pk}'")
                return redirect('index')

            location = BangViTriTau.objects.filter(
                Q(Ngay__gte=start_date) & Q(Ngay__lte=end_date),
                IDTau = ship
            ).order_by('Ngay').values('ViDo', 'KinhDo', 'Ngay')
            # print(location)
            for loc in location:
                loc['Ngay'] = loc['Ngay'].strftime('%Y-%m-%d %H:%M:%S')
            # print(location)

            location_data = list(location)

            info = {
                'SoDangKy': ship.SoDangKy,
                'ChuTau': ship.IDChuTau.HoTen,
                'ThuyenTruong': ship.IDThuyenTruong.HoTen,
                'ViDo': location_data[-1]['ViDo'],
                'KinhDo': location_data[-1]['KinhDo'],
                'Ngay': location_data[-1]['Ngay']
            }


            return JsonResponse({
                'message': f'Nhật ký hải trình tàu {ship.SoDangKy}',
                'status': 200,
                'location': location_data,
                'info': info,
            }, status=200)
        
        except json.JSONDecodeError:
            return JsonResponse({
                'message': 'Dữ liệu không hợp lệ',
                'status': 400
            }, status=400)
    else:
        return JsonResponse({
            'message': 'Method not allowed',
            'status': 405
        }, status=405)
    

# Lấy info của tất cả các tàu và hiển thị trên map 
@login_required(login_url='/login/')
def get_all_location_api(request):
    # ships = BangTau.objects.all()

    # Lấy danh sách tàu đã có vị trí
    ships = []
    if request.user.user_type == '1' or request.user.is_staff:
        ships_with_position = BangTau.objects.prefetch_related('bangvitritau_set')
    elif request.user.user_type == '2':
        ships_with_position = BangTau.objects.filter(CangCaDangKy=request.user.staff.cangca).prefetch_related('bangvitritau_set')
    else:
        return render(request, '403.html', {}, status=403)

    for ship in ships_with_position:
        if ship.bangvitritau_set.exists():
            ships.append(ship)

    info = []
    for i in ships:
        new_loc = i.bangvitritau_set.order_by('-Ngay').first()
        ship_info = {
            'SoDangKy': i.SoDangKy,
            'ChuTau': i.IDChuTau.HoTen,
            'ThuyenTruong': i.IDThuyenTruong.HoTen,
            'ViDo': new_loc.ViDo,
            'KinhDo': new_loc.KinhDo,
            'NgayCapNhat': new_loc.Ngay,
        }
        info.append(ship_info)
    return JsonResponse({
        'status': 200,
        'message': 'Cập nhật thông tin vị trí các tàu thành công!',
        'info': info,
    }, status=200)


# Báo cáo thống kê
@login_required(login_url='/login/')
def report_view(request):
    fishing_port = BangCangCa.objects.all()
    titles = ["STT", "Tên Loài Cá", "Sản Lượng"]

    # Lấy thời điểm 24h trước 
    time_threshold = timezone.now() - timedelta(hours=24)

    # Lấy danh sách các mẻ lưới có thời gian thu ngư cụ trong khoảng 24h
    if request.user.user_type == '1' or request.user.is_staff:
        bang_me_luoi_24h = BangMeLuoi.objects.filter(ThoiDiemThuNguCu__lte=time_threshold)
    elif request.user.user_type == '2':
        bang_me_luoi_24h = BangMeLuoi.objects.filter(ThoiDiemThuNguCu__lte=time_threshold)
    else:
        return render(request, '403.html', {}, status=403)
    # print(bang_me_luoi_24h)
    # Tính tổng sản lượng của mỗi loại cá được đánh bắt trong 24h
    items = (
        BangLoaiCaDuocDanhBatTrongMeLuoi.objects
        .filter(IDMeLuoi__in=bang_me_luoi_24h)
        .values('IDLoaiCa__Ten')
        .annotate(tong_san_luong=Sum('SanLuong'))
        .order_by('-tong_san_luong').all()
    )

    return render(request, 'core/report.html', {
        'fishing_port': fishing_port,
        'titles': titles,
        'items': items,
    }, status=200)


@login_required(login_url='/login/')
def search_report_view(request):
    query = request.GET.get('q')
    start_date = request.GET.get('start-date')
    end_date = request.GET.get('end-date')
    
    fishing_port = BangCangCa.objects.all()
    titles = ["STT", "Tên Loài Cá", "Sản Lượng"]

    try:
        gate = BangCangCa.objects.get(pk=query)
    except BangCangCa.DoesNotExist:
        messages.info(request, f"Không tìm thấy thông tin cảng cá với query = '{query}'")
        return redirect('report-view')

    # print(gate)

    # ngày về bến nằm trong khoảng start date và end date
    sea_trip_list = BangChuyenBien.objects.filter(Q(NgayVeBen__gte=start_date) & Q(NgayVeBen__lte=end_date) & Q(CangVeBen=gate))
    if len(sea_trip_list) == 0:
        messages.info(request, f"Không tìm thấy chuyến biển về cảng {gate.Ten} trong khoảng thời gian từ {start_date} đến {end_date}")
        return redirect('report-view')
    fish_totals = []

    for sea_trip in sea_trip_list:
        net_list = sea_trip.bangmeluoi_set.all()

        for net in net_list:
            fish_list = net.bangloaicaduocdanhbattrongmeluoi_set.values('IDLoaiCa__Ten').annotate(tong_san_luong=Sum('SanLuong'))

            fish_totals += fish_list
    fish_totals.sort(key=lambda x: x['tong_san_luong'], reverse=True)
    # print(fish_totals)
    return render(request, 'core/report-search.html', {
        'fishing_port': fishing_port,
        'titles': titles,
        'items': fish_totals,
        'selected_port': gate,
        'start_date': start_date,
        'end_date': end_date
    }, status=200) 


@login_required(login_url='/login/')
def search_top_10_fishing_api(request):
    if request.method == 'POST':
        data = json.loads(request.body)

        query = data['query']
        start_date = data['start_date']
        end_date = data['end_date']

        try:
            gate = BangCangCa.objects.get(pk=query)
        except BangCangCa.DoesNotExist:
            messages.info(request, f"Không tìm thấy thông tin cảng cá với query = '{query}'")
            return redirect('report-view')

        # ngày về bến nằm trong khoảng start date và end date
        sea_trip_list = BangChuyenBien.objects.filter(Q(NgayVeBen__gte=start_date) & Q(NgayVeBen__lte=end_date) & Q(CangVeBen=gate))
        if len(sea_trip_list) == 0:
            messages.info(request, f"Không tìm thấy chuyến biển về cảng {gate.Ten} trong khoảng thời gian từ {start_date} đến {end_date}")
            return redirect('report-view')
        fish_totals = []

        for sea_trip in sea_trip_list:
            net_list = sea_trip.bangmeluoi_set.all()

            for net in net_list:
                fish_list = net.bangloaicaduocdanhbattrongmeluoi_set.values('IDLoaiCa__Ten').annotate(tong_san_luong=Sum('SanLuong'))

                fish_totals += fish_list
        fish_totals.sort(key=lambda x: x['tong_san_luong'], reverse=True)
        
        bundles = []
        for item in fish_totals:
            data = {
                'fish_name': item['IDLoaiCa__Ten'],
                'qty': item['tong_san_luong']
            }
            bundles.append(data)
        
        return JsonResponse({
            'message': f'Top 10 loại cá được đánh bắt nhiều nhất trong khoảng thời gian từ {start_date} đến {end_date} ',
            'status': 200,
            'bundles': bundles[:10],
            'start_date': start_date,
            'end_date': end_date
        }, status=200)

    else:
        return JsonResponse({
            'message': 'method not allowed',
            'success': False,
        }, status=405)


# Lấy 10 loại cá được đánh bắt nhiều nhất phục vụ cho biểu đồ
@login_required(login_url='/login/')
def top_10_fishing_api(request):
    # Lấy thời điểm 24h trước 
    time_threshold = timezone.now() - timedelta(hours=24)

    # Lấy danh sách các mẻ lưới có thời gian thu ngư cụ trong khoảng 24h
    bang_me_luoi_24h = BangMeLuoi.objects.filter(ThoiDiemThuNguCu__lte=time_threshold)
    # print(bang_me_luoi_24h)
    # Tính tổng sản lượng của mỗi loại cá được đánh bắt trong 24h
    top_loai_ca = (
        BangLoaiCaDuocDanhBatTrongMeLuoi.objects
        .filter(IDMeLuoi__in=bang_me_luoi_24h)
        .values('IDLoaiCa__Ten')
        .annotate(tong_san_luong=Sum('SanLuong'))
        .order_by('-tong_san_luong')[:10]
    )
    # print(top_loai_ca)
    bundles = []
    for item in top_loai_ca:
        data = {
            'fish_name': item['IDLoaiCa__Ten'],
            'qty': item['tong_san_luong']
        }
        bundles.append(data)
    # print(bundles)
    return JsonResponse({
        'message': 'Top 10 loại cá được đánh bắt nhiều nhất trong 24h qua',
        'status': 200,
        'bundles': bundles
    }, status=200)


# Nhật ký khai thác thủy sản 
@login_required(login_url='/login/')
def journal_view(request):
    titles = ["STT", "Mã tàu", "Ngày tạo nhật ký", "Chủ tàu", "CMND (CCCD)", "Mã nhật ký", "Ghi chú", "Thao tác"]
    if request.user.user_type == '1' or request.user.is_staff:
        ships = BangTau.objects.all().order_by('SoDangKy')
        items = []
        for ship in ships:
            newest_journal = ship.bangnhatky_set.order_by('-NgayTao').first()
            if newest_journal is not None:
                # print(newest_journal)
                items.append(newest_journal)
        return render(request, 'core/journal.html', {
            'titles': titles,
            'items': items,
        }, status=200)
    else:
        ships = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca)).order_by('SoDangKy')
        items = []
        for ship in ships:
            newest_journal = ship.bangnhatky_set.order_by('-NgayTao').first()
            if newest_journal is not None:
                items.append(newest_journal)
        return render(request, 'core/journal.html', {
            'titles': titles,
            'items': items,
        }, status=200)



# pdf view 
@login_required(login_url='/login/')
def journal_pdf_view(request, pk):
    response = generate_journal_pdf(request, pk)
    return HttpResponse(response.content, content_type='application/pdf')


# pdf detail view 
@login_required(login_url='/login/')
def generate_journal_pdf(request, pk):
    try:
        ediary = BangNhatKy.objects.get(pk=pk)
    except BangNhatKy.DoesNotExist:
        messages.info(request, f"Không tìm thấy thông tin nhật ký với ID = '{pk}'") 
        return redirect('journal-view')
    except Exception as e:
        messages.info(request, f"Lỗi khi truy xuất thông tin nhật ký!!!")
        return redirect('journal-view')

    # Tạo một đối tượng HttpResponse với content_type là 'application/pdf'
    response = HttpResponse(content_type='application/pdf')
    # response['Content-Disposition'] = 'attachment; filename="example.pdf"'
    response['Content-Disposition'] = 'inline; filename="example.pdf"'

    page_width, page_height = landscape(letter)
    pdf = canvas.Canvas(
        response,
        pagesize=(page_width, page_height)
    )

    # Vẽ đường
    pdf.setStrokeColor(colors.black)
    pdf.setLineWidth(1)
    pdf.line(320, 515, 470, 515)  # Vẽ đường ngang trên cùng của Frame

    styles = getSampleStyleSheet()
    style_normal = styles['Normal']

    font_normal_path = 'times.ttf'
    font_bold_path = 'timesbd.ttf'

    # font_normal_path = '/home/administrator/eproject/times.ttf'
    # font_bold_path = '/home/administrator/eproject/timesbd.ttf'

    pdfmetrics.registerFont(TTFont('Times_New_Roman', font_normal_path))
    pdfmetrics.registerFont(TTFont('Times_New_Roman_Bold', font_bold_path))

    h2 = ParagraphStyle('CustomHeading2', parent=styles['Heading2'])
    h2.fontName = 'Times_New_Roman_Bold'
    h2.alignment = 1

    h2_normal = ParagraphStyle('CustomHeading3', parent=styles['Heading2'])
    h2_normal.fontName = "Times_New_Roman"
    h2_normal.alignment = 1

    # Vẽ hình chữ nhật
    # pdf.rect(515, 515, 200, 25)

    # Tính toán vị trí trung tâm của hình chữ nhật
    rect_x = 515
    rect_y = 515
    rect_width = 200
    rect_height = 25
    text = ediary.MaNhatKy
    text_width, text_height = pdf.stringWidth(text), 16
    text_x = rect_x + (rect_width - text_width) / 2
    text_y = rect_y + (rect_height - text_height) / 2

    # Vẽ hình chữ nhật
    pdf.rect(rect_x, rect_y, rect_width, rect_height)
    pdf.setFont("Times_New_Roman", 16)
    # Thêm văn bản vào trung tâm của hình chữ nhật
    pdf.drawString(text_x, text_y, text)

    frm1_width = page_width - (2.5 * cm) * 2
    frm1_height = (7/10) * (page_height - 2.5 *cm)

    flow_obj1 = []
    frame1 = Frame(
        x1=2.5*cm,
        y1=page_height - frm1_height - 2.5*cm,
        width=frm1_width,
        height=frm1_height,
        showBoundary=1
    )

    paragraph_normal = styles['Normal']
    paragraph_normal.fontSize = 13
    paragraph_normal.spaceAfter = 1.3 * paragraph_normal.fontSize 
    paragraph_normal.fontName = "Times_New_Roman"

    flow_obj1.append(Paragraph("TỔNG CỤC THỦY SẢN", h2))
    flow_obj1.append(Paragraph("NHẬT KÝ KHAI THÁC THỦY SẢN", h2))
    flow_obj1.append(Paragraph(f"NGHỀ CHÍNH: {ediary.IDThietBi.NgheChinh.Ten}.........(**)", h2_normal))

    flow_obj1.append(Paragraph(f"1. Họ và tên chủ tàu: {ediary.IDThietBi.IDChuTau.HoTen}.................; 2. Họ và tên thuyền trưởng: {ediary.IDThietBi.IDThuyenTruong.HoTen}..................", paragraph_normal))
    flow_obj1.append(Paragraph(f"3. Số đăng ký tàu: {ediary.IDThietBi.SoDangKy}.....; 4. Chiều dài lớn nhất của tàu: {ediary.IDThietBi.ChieuDaiLonNhat} m; 5. Tổng công suất máy chính: {ediary.IDThietBi.CongSuatMay} KW", paragraph_normal))
    flow_obj1.append(Paragraph(f"6. Số Giấy phép khai thác thủy sản: GP-120.................Thời hạn đến: 27/12/2039...........................................", paragraph_normal))
    flow_obj1.append(Paragraph(f"7. Nghề phụ 1: {ediary.IDThietBi.NghePhu1.Ten}..; 8. Nghề phụ 2: Đánh bắt cá ngừ.....................................", paragraph_normal))
    flow_obj1.append(Paragraph(f"9. Kích thước chủ yếu của ngư cụ (ghi cụ thể theo nghề chính): ", paragraph_normal))
    flow_obj1.append(Paragraph(f"a) Nghề câu: Chiều dài toàn bộ vàng câu: ................m; Số lưỡi câu: ..................................................lưỡi", paragraph_normal))
    flow_obj1.append(Paragraph(f"b) Nghề lưới vây, rê: Chiều dài toàn bộ lưới: ...........m; Chiều cao lưới: ..................................................m", paragraph_normal))
    flow_obj1.append(Paragraph(f"c) Nghề lưới chụp: Chu vi miệng lưới: ...................m; Chiều cao lưới: ..................................................m", paragraph_normal))
    flow_obj1.append(Paragraph(f"d) Nghề lưới kéo: Chiều dài giềng phao: .................m; Chiều dài toàn bộ lưới: ..........................................m", paragraph_normal))
    flow_obj1.append(Paragraph(f"e) Nghề khác:...........................................................", paragraph_normal))

    frame1.addFromList(flow_obj1, pdf)

    frm2_width = (1/4)*(page_width - 5 * cm)
    frm2_height = (1/6)  * (page_height - 2.5 *cm)
    flow_obj2 = []
    frame2 = Frame(
        x1=2.5*cm,
        y1=page_height - (frm1_height + frm2_height) - 2.5*cm,
        width=frm2_width,
        height=frm2_height,
        showBoundary=1
    )

    paragrap_bold = styles['Title']
    paragrap_bold.fontSize = 13
    paragrap_bold.spaceAfter = 1.3 * paragrap_bold.fontSize 
    paragrap_bold.fontName = "Times_New_Roman_Bold"
    paragrap_bold.alignment = 0

    flow_obj2.append(Paragraph(f"Chuyến biển số: {ediary.IDChuyenBien.ChuyenBienSo}...........", paragrap_bold))


    frame2.addFromList(flow_obj2, pdf)

    frm3_width = (3/4)*(page_width - 5 * cm)
    frm3_height = (1/6) * (page_height - 2.5 *cm)
    flow_obj3 = []
    frame3 = Frame(
        x1=2.5*cm + frm2_width,
        y1=page_height - (frm1_height + frm3_height) - 2.5*cm,
        width=frm3_width,
        height=frm3_height,
        showBoundary=1
    )

    flow_obj3.append(Paragraph(f"10. Cảng đi: {ediary.IDChuyenBien.CangXuatBen.Ten}.......; Thời gian đi: Ngày {ediary.IDChuyenBien.NgayXuatBen.day}... tháng {ediary.IDChuyenBien.NgayXuatBen.month}... năm {ediary.IDChuyenBien.NgayXuatBen.year}...",paragraph_normal))
    flow_obj3.append(Paragraph(f"11. Cảng về: {ediary.IDChuyenBien.CangVeBen.Ten}.......; Thời gian cập cảng: Ngày {ediary.IDChuyenBien.NgayVeBen.day}... tháng {ediary.IDChuyenBien.NgayVeBen.month}... năm {ediary.IDChuyenBien.NgayVeBen.year}...",paragraph_normal))
    flow_obj3.append(Paragraph(f"12. Nộp Nhật ký: Ngày ..... tháng ..... năm .......; Vào Sổ số: {ediary.MaNhatKy}",paragraph_normal))

    frame3.addFromList(flow_obj3, pdf)

    pdf.showPage()

    h2_left = ParagraphStyle('CustomHeading2', parent=styles['Heading2'])
    h2_left.fontName = 'Times_New_Roman_Bold'
    h2_left.alignment = 0


    flow_obj4 = []
    frame4 = Frame(
        x1=2.5*cm,
        y1=2.5*cm,
        width=(page_width - 5*cm),
        height=(page_height - 4*cm),
        showBoundary=0
    )
    flow_obj4.append(Paragraph("I. THÔNG TIN VỀ HOẠT ĐỘNG KHAI THÁC THỦY SẢN", h2_left))
    flow_obj4.append(Paragraph("1.Thông tin mẻ lưới/câu", paragrap_bold))

    ratio = (page_width - 4*cm) / 40
    colWidths = []
    colWidths.append(ratio*2)
    colWidths.append(ratio*4)
    colWidths.append(ratio*3)
    colWidths.append(ratio*3)
    colWidths.append(ratio*4)
    for i in range(8):
        colWidths.append(ratio*3)
    # print(colWidths)

    sea_trip = ediary.IDChuyenBien
    net_list = BangMeLuoi.objects.filter(IDChuyenBien=sea_trip).order_by('MeLuoiSo')[:3]
    sum_qty = []
    if net_list is not None:
        for net in net_list:
            sum = [0] * 6
            fish_list = BangLoaiCaDuocDanhBatTrongMeLuoi.objects.filter(IDMeLuoi=net).order_by('-SanLuong')[:6]
            # print(fish_list)
            
            if fish_list is not None:
                for index, fish in enumerate(fish_list):
                    sum[index] = fish.SanLuong
            sum_qty.append(sum)
    print(sum_qty)
    tmp_list = []
    for i, item in enumerate(net_list):
        date_start = item.ThoiDiemThaNguCu
        date_end = item.ThoiDiemThuNguCu
        # print(item)
        tmp_ = [i+1, f"{date_start.day}-{date_start.month}-{date_start.year}", f"{str(item.ViDoThaNguCu)[0:7]}", f"{str(item.KinhDoThaNguCu)[0:7]}", f"{date_end.day}-{date_end.month}-{date_end.year}", f"{str(item.ViDoThuNguCu)[0:7]}", f"{str(item.KinhDoThuNguCu)[0:7]}", f"{sum_qty[i][0]}", f"{sum_qty[i][1]}", f"{sum_qty[i][2]}", f"{sum_qty[i][3]}", f"{sum_qty[i][4]}", f"{sum_qty[i][5]}", item.TongSanLuong]
        tmp_list.append(tmp_)
    # print(tmp_list)

    data = [
        ['Mẻ\nthứ', 'Thời \nđiểm bắt\nđầu thả', 'Vị trí thả', '', 'Thời\nđiểm kết\nthúc thu', 'Vị trí thu', '', 'Sản lượng các loài chủ yếu**(kg)', '', '', '', '', '', 'Tổng \nsản\nlượng'],
        ['', '', 'Vĩ độ', 'Kinh độ', '', 'Vĩ độ', 'Kinh độ', 'Loài 1', 'Loài 2', 'Loài 3', 'Loài 4', 'Loài 5', 'Loài 6', ''],
        ['', '', '', '', ''],
    ]
    # [1, '20/01/2023', 13.04, 109.01, '20/02/2023', 13.05, 110.11, 100, 200, 3000, 400, 500, 900, sum_qty[0]],
    # [2, '21/02/2023', 13.04, 109.01, '21/03/2023', 13.05, 110.11, 100, 200, 3000, 400, 500, 900, sum_qty[1]],
    # [3, '21/02/2023', 13.04, 109.01, '21/03/2023', 13.05, 110.11, 100, 200, 3000, 400, 500, 900, sum_qty[2]],
    data.extend(tmp_list)

    table = Table(data, colWidths=colWidths)
    tstyle = TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONT", (0, 0), (-1, -1), "Times_New_Roman",13),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ("SPAN", (0, 0), (0, 2)),  # Row span from row 0 to row 1 for the first column (col - row) (col - row )  (Row span thì col giữ nguyên) # mẻ thứ
        ("SPAN", (1, 0), (1, 2)),   # Row span from row 1 to row 2 for the second column  # thời điểm bắt đầu thả
        ("SPAN", (2, 0), (3, 0)),  # Column span from column 0 to column 2 for the first Row (col span thì row giữ nguyên) # vị trí thả
        ("SPAN", (2, 1), (2, 2)), # vĩ độ thả
        ("SPAN", (3, 1), (3, 2)), # kinh độ thả
        ("SPAN", (4, 0), (4, 2)), # thời điểm kết thúc thu 
        ("SPAN", (5, 0), (6, 0)), # vị trí thu
        ("SPAN", (5, 1), (5, 2)), # vĩ độ thu
        ("SPAN", (6, 1), (6, 2)), # kinh độ thu
        ("SPAN", (7, 0), (12, 0)), # sản lượng các loài
        ("SPAN", (7, 1), (7, 2)), # loài
        ("SPAN", (8, 1), (8, 2)), # loài
        ("SPAN", (9, 1), (9, 2)), # loài
        ("SPAN", (10, 1), (10, 2)), # loài
        ("SPAN", (11, 1), (11, 2)), # loài
        ("SPAN", (12, 1), (12, 2)), # loài
        ("SPAN", (13, 0), (13, 2)) # tổng sản lượng
    ])
    table.setStyle(tstyle)
    flow_obj4.append(table)
    flow_obj4.append(Spacer(1, 0.25*cm))
    flow_obj4.append(Paragraph("**Ghi các đối tượng khai thác chính theo từng nghề (Kéo, Rê, Vây, Câu, Chụp…). Đối với các nghề khai thác cá ngừ cần ghi rõ sản lượng của từng loài như: cá ngừ Vây vàng, cá ngừ Mắt to, cá ngừ Vằn (Sọc dưa), cá ngừ khác (Chù, ồ…).", paragraph_normal))
    flow_obj4.append(Paragraph("2.Thông tin về các loài nguy cấp quý hiếm", paragrap_bold))
    flow_obj4.append(Paragraph("Cá voi/Cá heo/Bò biển/Quản đồng/Vích/Đồi mồi dứa/Đồi mồi/Rùa da/Loài khác (Ghi tên cụ thể)", paragraph_normal))

    data2 = [
        ["Mẻ","Loài","Thời điểm\nbắt gặp", "Khối\nlượng/con\n(kg)", "Số lượng\nước tính", "Kích thước\nước tính\n(cm)", "Bắt gặp trong quá trình khai thác\n(chọn 1)", "", "", "Tình trạng bắt gặp (chọn 1)"],
        ["", "", "", "", "", "", "Lưới/câu", "Kéo lưới", "Khác", "Sống", "Chết", "Bị thương"],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    colWidths2 = [1*cm, 2.2*cm, 3*cm]
    for i in range(9):
        colWidths2.append(2.2*cm)
    table2 = Table(data2, colWidths=colWidths2)
    tstyle2 = TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONT", (0, 0), (-1, -1), "Times_New_Roman",13),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ("SPAN", (0, 0), (0, 1)), # mẻ
        ("SPAN", (1, 0), (1, 1)), # loài
        ("SPAN", (2, 0), (2, 1)), # thời điểm bắt gặp
        ("SPAN", (3, 0), (3, 1)), # khối lượng/con (kg)
        ("SPAN", (4, 0), (4, 1)), # số lượng
        ("SPAN", (5, 0), (5, 1)), # kích thước
        ("SPAN", (6, 0), (8, 0)), # bắt gặp trong quá trình khai thác
        ("SPAN", (9, 0), (11, 0)), # tình trạng bắt gặp
    ])
    table2.setStyle(tstyle2)
    flow_obj4.append(table2)
    flow_obj4.append(Spacer(1, 0.25*cm))
    flow_obj4.append(Paragraph("Thông tin bổ sung về loài (nếu có): (Về màu sắc loài; thiết bị, thẻ gắn số trên cá thể;…và các thông tin khác nếu có)\n\
        ……………………………………………………………………………………………………………………………………………………\
    …………………………..………………………………………………………………………………………………………………………\
    ……………………………………………………………………………………….", style_normal))
    frame4.addFromList(flow_obj4, pdf)
    pdf.showPage()

    flow_obj5 = []
    frame5 = Frame(
        x1=2.5*cm,
        y1=2.5*cm + 0.4*(page_height - 5*cm),
        width=(page_width - 5*cm),
        height=0.6*(page_height - 3*cm),
        showBoundary=0
    )

    flow_obj5.append(Paragraph("II. THÔNG TIN VỀ HOẠT ĐỘNG CHUYỂN TẢI (nếu có)", h2_left))
    data3 = [
        ["TT", "Ngày, tháng", "Thông tin tàu thu\nmua/chuyển tải", "", "Vị trí thu mua,\nchuyển tải", "", "Đã bán/chuyển tải", "", "Thuyền trưởng \ntàu thu mua,\nchuyểntải"],
        ["", "", "Số đăng ký\ntàu", "Số giấy phép\nkhaithác", "Vĩ độ", "Kinh độ", "Tên loài\nthủy sản", "Khối lượng\n(kg)", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
    ]
    table3 = Table(data3, colWidths=[1*cm, 3*cm, 2.5*cm, 3*cm, 2*cm, 2*cm, 3*cm, 3*cm, 3.5*cm])
    tsyle3 = TableStyle([
        # col, row
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONT", (0, 0), (-1, -1), "Times_New_Roman",13),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ("SPAN", (0, 0), (0, 1)), # tt
        ("SPAN", (1, 0), (1, 1)), # ngày, tháng
        ("SPAN", (2, 0), (3, 0)), # thông tin thu mua, chuyển tải
        ("SPAN", (4, 0), (5, 0)), # vị trí thu mua, chuyển tải
        ("SPAN", (6, 0), (7, 0)), # đã bán, chuyển tải
        ("SPAN", (8, 0), (8, 1)), # thuyển trưởng ký
    ])
    table3.setStyle(tsyle3)
    flow_obj5.append(table3)
    flow_obj5.append(Spacer(1, 0.5*cm))
    # flow_obj5.append(Paragraph("Ngày ...... tháng ...... năm ......", paragraph_normal))
    # flow_obj5.append(Paragraph("Thuyền trưởng", paragrap_bold))
    # flow_obj5.append(Paragraph("(ký, ghi rõ họ tên)", paragraph_normal))

    frame5.addFromList(flow_obj5, pdf)

    flow_obj6 = []
    frame6 = Frame(
        x1=16*cm,
        y1=4*cm,
        width=10*cm,
        height=6*cm,
        showBoundary=0
    )

    paragraph_bold_center = styles['Title']
    paragraph_bold_center.fontSize = 13
    paragraph_bold_center.alignment = 1
    paragraph_bold_center.fontName = "Times_New_Roman_Bold"
    paragraph_bold_center.spaceAfter = 1.1 * paragraph_bold_center.fontSize 

    paragraph_normal_center = styles['Normal']
    paragraph_normal_center.fontSize = 13
    paragraph_normal_center.alignment = 1
    paragraph_normal_center.fontName = "Times_New_Roman"
    paragraph_normal_center.spaceAfter = 1.1 * paragraph_normal_center.fontSize 

    flow_obj6.append(Paragraph("Ngày ..... tháng ..... năm ......", paragraph_normal_center))
    flow_obj6.append(Paragraph("Thuyền trưởng", paragraph_bold_center))
    flow_obj6.append(Paragraph("(ký, ghi rõ họ tên)", paragraph_normal_center))
    frame6.addFromList(flow_obj6, pdf)

    pdf.showPage()

    pdf.save()

    return response 


@login_required(login_url='/login/')
def search_journal_view(request):
    titles = ["STT", "Mã tàu", "Ngày tạo nhật ký", "Chủ tàu", "CMND (CCCD)", "Mã nhật ký", "Ghi chú", "Thao tác"]
    if request.user.user_type == '1' or request.user.is_staff:
        query = request.GET.get('q')
        query_type = request.GET.get('query-type')
        
        # mã tàu
        if query_type == '1':
            ships = BangTau.objects.filter(Q(SoDangKy__icontains=query))
            items = []
            for ship in ships:
                journal_list = ship.bangnhatky_set.all()
                items.extend(journal_list)
            if len(items) == 0:
                messages.info(request, f"Không tìm thấy thông tin nhật ký liên quan đến tàu có số đăng ký query = '{query}'")

        # mã chủ tàu
        if query_type == '2':
            pass 

        # mã thuyền trưởng
        if query_type == '3':
            pass 

        # mã nhật ký 
        if query_type == '4':
            items = BangNhatKy.objects.filter(Q(MaNhatKy__icontains=query))
            if len(items) == 0:
                messages.info(request, f"Không tìm thấy thông tin nhật ký với mã query = '{query}'")

        # mã chuyến biển
        if query_type == '5':
            tmp = BangChuyenBien.objects.filter(Q(IDChuyenBien__icontains=query) | Q(ChuyenBienSo__icontains=query))
            items = []
            for i in tmp:
                journal_list = i.bangnhatky_set.all()
                items.extend(journal_list) 
            print(len(items))
            if len(items) == 0:
                messages.info(request, f"Không tìm thấy thông tin nhật ký liên quan đến chuyển biến với query = '{query}'")  

        return render(request, 'core/journal.html', {
            'titles': titles,
            'items': items,
        }, status=200)

    elif request.user.user_type == '2':
        pass
    else:
        return render(request, '403.html', {}, status=403)


# quản lý tàu cá 
@login_required(login_url='/login/')
def device_view(request):
    '''Quản lý tàu cá'''
    titles = ["STT","IMO","Số hiệu tàu","Chủ tàu","Loại tàu","Nơi đăng ký","Thuyền trưởng","Cảng cá ĐK","TBNKKT","Thao tác"]
    
    if request.user.user_type == '1' or request.user.is_staff:
        items = BangTau.objects.all().order_by('SoDangKy')
    elif request.user.user_type == '2':
        items = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca)).order_by('SoDangKy')
    else:
        return render(request, '403.html', {}, status=403)

    return render(request, 'core/device.html', {
        'titles': titles,
        'items': items, 
    }, status=200)


# thêm thông tin tàu cá mới
@login_required(login_url='/login/')
def add_new_device_view(request):
    '''Thêm tàu cá mới'''
    if request.user.user_type == '1' or request.user.is_staff:
        if request.method == 'POST':
            # print(request.POST)

            TenTau = request.POST.get('tenTau')
            try:
                ChuTau = BangChuTau.objects.get(pk=request.POST['chuTau'])
            except Exception as e:
                return redirect('add-device')

            try:
                LoaiTau = BangMaLoaiTau.objects.get(pk=request.POST['loaiTau'])
            except Exception as e:
                return redirect('add-device')

            # xử lý số đăng ký trùng, IMO trùng
            SoDangKy = request.POST.get('soDangKy')
            HoHieu = request.POST.get('hoHieu')
            CoHieu = request.POST.get('coHieu')
            IMO = request.POST.get('IMO')
            
            try:
                NoiDangKy = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['noiDangKy'])
            except Exception as e:
                return redirect('add-device')
            print(NoiDangKy)
            try:
                CangCaDangKy = BangCangCa.objects.get(pk=request.POST['cangCaDangKy'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                CangCaPhu = BangCangCa.objects.get(pk=request.POST['cangCaPhu'])
            except:
                return redirect('add-device')

            try:
                NgheChinh = BangNganhNgheKhaiThac.objects.get(pk=request.POST['ngheChinh'])
            except Exception as e:
                return redirect('add-device')
            
            NghePhu1 = BangNganhNgheKhaiThac.objects.get(pk=request.POST['nghePhu1'])
            NgayDangKy = request.POST.get('ngayDangKy')
            NgayHetHanDangKy = request.POST.get('ngayHetHanDangKy')
            NgaySanXuatTau = request.POST.get('ngaySanXuatTau')
            NgayHetHanSuDung = request.POST.get('ngayHetHanSuDung')
            
            try:
                MaThietBi = BangThietBiNhatKyKhaiThac.objects.get(pk=request.POST['maThietBi'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                ThuyenTruong = BangThuyenTruong.objects.get(pk=request.POST['thuyenTruong'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                Tinh = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['tinhThanhPho'])
            except Exception as e:
                return redirect('add-device')
            
            SoLuongThuyenVien = request.POST.get('soLuongThuyenVien')
            TongTaiTrong = request.POST.get('tongTaiTrong')
            ChieuDaiLonNhat = request.POST.get('chieuDaiLonNhat')
            ChieuRongLonNhat = request.POST.get('chieuRongLonNhat')
            CongSuatMay = request.POST.get('congSuatMay')
            MonNuoc = request.POST.get('monNuoc')
            DungTichHamCa = request.POST.get('dungTichHamCa')
            VanTocDanhBat = request.POST.get('vanTocDanhBat')
            VanTocHanhTrinh = request.POST.get('vanTocHanhTrinh')
            ThongSoNguCu = request.POST.get('thongSoNguCu')

            new_ship = BangTau.objects.create(
                SoDangKy=SoDangKy,
                TenTau=TenTau,
                HoHieu=HoHieu,
                CoHieu=CoHieu,
                IMO=IMO,
                NoiDangKy=NoiDangKy,
                CangCaDangKy=CangCaDangKy,
                NgheChinh=NgheChinh,
                NghePhu1=NghePhu1,
                NgayDangKy=NgayDangKy,
                NgayHetHanDangKy=NgayHetHanDangKy,
                TongTaiTrong=TongTaiTrong,
                ChieuDaiLonNhat=ChieuDaiLonNhat,
                ChieuRongLonNhat=ChieuRongLonNhat,
                CongSuatMay=CongSuatMay,
                MonNuoc=MonNuoc,
                SoThuyenVien=SoLuongThuyenVien,
                NgaySanXuat=NgaySanXuatTau,
                NgayHetHan=NgayHetHanSuDung,
                LoaiTau=LoaiTau,
                DungTichHamCa=DungTichHamCa,
                VanTocDanhBat=VanTocDanhBat,
                VanTocHanhTrinh=VanTocHanhTrinh,
                IDDevice=MaThietBi,
                IDChuTau=ChuTau,
                IDThuyenTruong=ThuyenTruong,
                IDTinh=Tinh,
                ThongSoNguCu=ThongSoNguCu
            )

            MaThietBi.is_active = True
            MaThietBi.save()
            ThuyenTruong.hasShip = True
            ThuyenTruong.save()

            messages.success(request, f"Thêm thông tin tàu {SoDangKy} thành công!!")
            return redirect('device-view')


        city_list = BangDonViHanhChinhCapTinh.objects.all()
        shipowners = BangChuTau.objects.all().order_by('HoTen')
        ship_type_list = BangMaLoaiTau.objects.all()
        gate_list = BangCangCa.objects.all().order_by('Ten')
        job_list = BangNganhNgheKhaiThac.objects.all()
        device_list = BangThietBiNhatKyKhaiThac.objects.all().order_by('SerialNumber').exclude(is_active=True) # danh sách thiết bị chưa kích hoạt
        captain_list = BangThuyenTruong.objects.all().order_by('HoTen').exclude(hasShip=True) # danh sách các thuyền trưởng chưa có tàu nào
        today = date.today()
        return render(request, 'core/add-new-device.html', {
            'shipowners': shipowners,
            'ship_type_list': ship_type_list,
            'city_list': city_list,
            'gate_list': gate_list,
            'job_list': job_list,
            'device_list': device_list,
            'captain_list': captain_list,
            'today': today,
        }, status=200)
    elif request.user.user_type == '2':
        if request.method == 'POST':
            TenTau = request.POST.get('tenTau')
            try:
                ChuTau = BangChuTau.objects.get(pk=request.POST['chuTau'])
            except Exception as e:
                return redirect('add-device')

            try:
                LoaiTau = BangMaLoaiTau.objects.get(pk=request.POST['loaiTau'])
            except Exception as e:
                return redirect('add-device')

            # xử lý số đăng ký trùng, IMO trùng
            SoDangKy = request.POST.get('soDangKy')
            HoHieu = request.POST.get('hoHieu')
            CoHieu = request.POST.get('coHieu')
            IMO = request.POST.get('IMO')
            
            try:
                NoiDangKy = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['noiDangKy'])
            except Exception as e:
                return redirect('add-device')
            print(NoiDangKy)
            try:
                CangCaDangKy = BangCangCa.objects.get(pk=request.POST['cangCaDangKy'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                CangCaPhu = BangCangCa.objects.get(pk=request.POST['cangCaPhu'])
            except:
                return redirect('add-device')

            try:
                NgheChinh = BangNganhNgheKhaiThac.objects.get(pk=request.POST['ngheChinh'])
            except Exception as e:
                return redirect('add-device')
            
            NghePhu1 = BangNganhNgheKhaiThac.objects.get(pk=request.POST['nghePhu1'])
            NgayDangKy = request.POST.get('ngayDangKy')
            NgayHetHanDangKy = request.POST.get('ngayHetHanDangKy')
            NgaySanXuatTau = request.POST.get('ngaySanXuatTau')
            NgayHetHanSuDung = request.POST.get('ngayHetHanSuDung')
            
            try:
                MaThietBi = BangThietBiNhatKyKhaiThac.objects.get(pk=request.POST['maThietBi'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                ThuyenTruong = BangThuyenTruong.objects.get(pk=request.POST['thuyenTruong'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                Tinh = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['tinhThanhPho'])
            except Exception as e:
                return redirect('add-device')
            
            SoLuongThuyenVien = request.POST.get('soLuongThuyenVien')
            TongTaiTrong = request.POST.get('tongTaiTrong')
            ChieuDaiLonNhat = request.POST.get('chieuDaiLonNhat')
            ChieuRongLonNhat = request.POST.get('chieuRongLonNhat')
            CongSuatMay = request.POST.get('congSuatMay')
            MonNuoc = request.POST.get('monNuoc')
            DungTichHamCa = request.POST.get('dungTichHamCa')
            VanTocDanhBat = request.POST.get('vanTocDanhBat')
            VanTocHanhTrinh = request.POST.get('vanTocHanhTrinh')
            ThongSoNguCu = request.POST.get('thongSoNguCu')

            new_ship = BangTau.objects.create(
                SoDangKy=SoDangKy,
                TenTau=TenTau,
                HoHieu=HoHieu,
                CoHieu=CoHieu,
                IMO=IMO,
                NoiDangKy=NoiDangKy,
                CangCaDangKy=CangCaDangKy,
                NgheChinh=NgheChinh,
                NghePhu1=NghePhu1,
                NgayDangKy=NgayDangKy,
                NgayHetHanDangKy=NgayHetHanDangKy,
                TongTaiTrong=TongTaiTrong,
                ChieuDaiLonNhat=ChieuDaiLonNhat,
                ChieuRongLonNhat=ChieuRongLonNhat,
                CongSuatMay=CongSuatMay,
                MonNuoc=MonNuoc,
                SoThuyenVien=SoLuongThuyenVien,
                NgaySanXuat=NgaySanXuatTau,
                NgayHetHan=NgayHetHanSuDung,
                LoaiTau=LoaiTau,
                DungTichHamCa=DungTichHamCa,
                VanTocDanhBat=VanTocDanhBat,
                VanTocHanhTrinh=VanTocHanhTrinh,
                IDDevice=MaThietBi,
                IDChuTau=ChuTau,
                IDThuyenTruong=ThuyenTruong,
                IDTinh=Tinh,
                ThongSoNguCu=ThongSoNguCu
            )

            MaThietBi.is_active = True
            MaThietBi.save()
            ThuyenTruong.hasShip = True
            ThuyenTruong.save()

            messages.success(request, f"Thêm thông tin tàu {SoDangKy} thành công!!")
            return redirect('device-view')

        city_list = BangDonViHanhChinhCapTinh.objects.all()
        shipowners = BangChuTau.objects.all().order_by('HoTen')
        ship_type_list = BangMaLoaiTau.objects.all()
        gate_list = BangCangCa.objects.all().order_by('Ten')
        job_list = BangNganhNgheKhaiThac.objects.all()
        device_list = BangThietBiNhatKyKhaiThac.objects.all().order_by('SerialNumber').exclude(is_active=True) # danh sách thiết bị chưa kích hoạt
        captain_list = BangThuyenTruong.objects.all().order_by('HoTen').exclude(hasShip=True) # danh sách các thuyền trưởng chưa có tàu nào
        today = date.today()
        return render(request, 'core/add-new-device.html', {
            'shipowners': shipowners,
            'ship_type_list': ship_type_list,
            'city_list': city_list,
            'gate_list': gate_list,
            'job_list': job_list,
            'device_list': device_list,
            'captain_list': captain_list,
            'today': today,
        }, status=200)
    else:
        return render(request, '403.html', {}, status=403)


# sửa thông tin tàu cá
@login_required(login_url='/login/')
def edit_device_view(request, pk):
    if request.user.user_type == '2' or request.user.user_type == '1' or request.user.is_staff:
        if request.method == 'POST':
            TenTau = request.POST.get('tenTau')
            try:
                ChuTau = BangChuTau.objects.get(pk=request.POST['chuTau'])
            except Exception as e:
                return redirect('add-device')

            try:
                LoaiTau = BangMaLoaiTau.objects.get(pk=request.POST['loaiTau'])
            except Exception as e:
                return redirect('add-device')

            # xử lý số đăng ký trùng, IMO trùng
            SoDangKy = request.POST.get('soDangKy')
            HoHieu = request.POST.get('hoHieu')
            CoHieu = request.POST.get('coHieu')
            IMO = request.POST.get('IMO')
            
            try:
                NoiDangKy = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['noiDangKy'])
            except Exception as e:
                return redirect('add-device')
            print(NoiDangKy)
            try:
                CangCaDangKy = BangCangCa.objects.get(pk=request.POST['cangCaDangKy'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                CangCaPhu = BangCangCa.objects.get(pk=request.POST['cangCaPhu'])
            except:
                return redirect('add-device')

            try:
                NgheChinh = BangNganhNgheKhaiThac.objects.get(pk=request.POST['ngheChinh'])
            except Exception as e:
                return redirect('add-device')
            
            NghePhu1 = BangNganhNgheKhaiThac.objects.get(pk=request.POST['nghePhu1'])
            NgayDangKy = request.POST.get('ngayDangKy')
            NgayHetHanDangKy = request.POST.get('ngayHetHanDangKy')
            NgaySanXuatTau = request.POST.get('ngaySanXuatTau')
            NgayHetHanSuDung = request.POST.get('ngayHetHanSuDung')
            
            try:
                MaThietBi = BangThietBiNhatKyKhaiThac.objects.get(pk=request.POST['maThietBi'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                ThuyenTruong = BangThuyenTruong.objects.get(pk=request.POST['thuyenTruong'])
            except Exception as e:
                return redirect('add-device')
            
            try:
                Tinh = BangDonViHanhChinhCapTinh.objects.get(pk=request.POST['tinhThanhPho'])
            except Exception as e:
                return redirect('add-device')
            
            SoLuongThuyenVien = request.POST.get('soLuongThuyenVien')
            TongTaiTrong = request.POST.get('tongTaiTrong')
            ChieuDaiLonNhat = request.POST.get('chieuDaiLonNhat')
            ChieuRongLonNhat = request.POST.get('chieuRongLonNhat')
            CongSuatMay = request.POST.get('congSuatMay')
            MonNuoc = request.POST.get('monNuoc')
            DungTichHamCa = request.POST.get('dungTichHamCa')
            VanTocDanhBat = request.POST.get('vanTocDanhBat')
            VanTocHanhTrinh = request.POST.get('vanTocHanhTrinh')
            ThongSoNguCu = request.POST.get('thongSoNguCu')

            ship = BangTau.objects.get(pk=pk)

            tbnkkt = ship.IDDevice
            tbnkkt.is_active = False
            tbnkkt.save()

            captain = ship.IDThuyenTruong
            captain.hasShip = False
            captain.save()

            ship.SoDangKy=SoDangKy
            ship.TenTau=TenTau
            ship.HoHieu=HoHieu
            ship.CoHieu=CoHieu
            ship.IMO=IMO
            ship.NoiDangKy=NoiDangKy
            ship.CangCaDangKy = CangCaDangKy
            ship.NgheChinh=NgheChinh
            ship.NghePhu1=NghePhu1
            ship.NgayDangKy=NgayDangKy
            ship.NgayHetHanDangKy=NgayHetHanDangKy
            ship.TongTaiTrong=TongTaiTrong
            ship.ChieuDaiLonNhat=ChieuDaiLonNhat
            ship.ChieuRongLonNhat=ChieuRongLonNhat
            ship.CongSuatMay=CongSuatMay
            ship.MonNuoc=MonNuoc
            ship.SoThuyenVien=SoLuongThuyenVien
            ship.NgaySanXuat=NgaySanXuatTau
            ship.NgayHetHan=NgayHetHanSuDung
            ship.LoaiTau=LoaiTau
            ship.DungTichHamCa=DungTichHamCa
            ship.VanTocDanhBat=VanTocDanhBat
            ship.VanTocHanhTrinh=VanTocHanhTrinh
            ship.IDDevice=MaThietBi
            ship.IDChuTau=ChuTau
            ship.IDThuyenTruong=ThuyenTruong
            ship.IDTinh=Tinh
            ship.ThongSoNguCu=ThongSoNguCu

            MaThietBi.is_active = True
            MaThietBi.save()
            ThuyenTruong.hasShip = True
            ThuyenTruong.save()

            ship.save()    
            messages.success(request, f"Cập nhật thông tin tàu {ship.SoDangKy} thành công!!")
            return redirect('device-view')


        try:
            item = BangTau.objects.get(pk=pk)
        except Exception as e:
            messages.info(request, f'Không tìm thấy thông tin tàu hợp lệ!')
            redirect('device-view')
        
        shipowners = BangChuTau.objects.all().order_by('HoTen')
        ship_type_list = BangMaLoaiTau.objects.all()
        city_list = BangDonViHanhChinhCapTinh.objects.all()
        gate_list = BangCangCa.objects.all().order_by('Ten')
        job_list = BangNganhNgheKhaiThac.objects.all()
        device_list = BangThietBiNhatKyKhaiThac.objects.all().order_by('SerialNumber').exclude(is_active=True)
        captain_list = BangThuyenTruong.objects.all().order_by('HoTen').exclude(hasShip=True)

        return render(request, 'core/edit-device.html', {
            'item': item,
            'shipowners': shipowners,
            'ship_type_list': ship_type_list,
            'city_list': city_list,
            'gate_list': gate_list,
            'job_list': job_list,
            'device_list': device_list,
            'captain_list': captain_list,
        }, status=200)
    else:
        return render(request, '403.html', {}, status=403) 

# xóa thông tin tàu cá
@login_required(login_url='/login/')
def delete_device_view(request, pk):
    if request.user.user_type == '1' or request.user.is_staff or request.user.user_type == '2':
        try:
            ship = BangTau.objects.get(pk=pk)
            soDangKy = ship.SoDangKy

            tbnkkt = ship.IDDevice
            tbnkkt.is_active = False
            tbnkkt.save()

            captain = ship.IDThuyenTruong
            captain.hasShip = False
            captain.save()
            
            ship.delete()
            messages.success(request, f"Xóa tàu {soDangKy} thành công!!")
            return redirect('device-view')
        except Exception as e:
            messages.error(request, f"Tàu mới pk = {pk} không tồn tại!!")
            return redirect('device-view')
    else:
        return render(request, '403.html', {}, status=403) 


# Tìm kiếm thông tin tàu cá
@login_required(login_url='/login/')
def search_device_view(request):
    query = request.GET.get('q')
    query_type = request.GET.get('query-type')
    titles = ["STT","IMO","Số hiệu tàu","Chủ tàu","Loại tàu","Nơi đăng ký","Thuyền trưởng","Cảng cá ĐK","TBNKKT","Thao tác"]
    if request.user.user_type == '1' or request.user.is_staff:
        items = BangTau.objects.all().order_by('-ID')

        if query_type == '1':
            items = BangTau.objects.filter(Q(IMO__icontains=query)).order_by('-ID')
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy IMO hợp lệ!')
        elif query_type == '2':
            items = BangTau.objects.filter(Q(SoDangKy__icontains=query)).order_by('-ID') 
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy số đăng ký tàu hợp lệ!')
        elif query_type == '3':
            users = BangChuTau.objects.filter(Q(HoTen__icontains=query)).order_by('-ID')
            # print(users)
            items = []
            for user in users:
                item = user.bangtau_chutau.all()
                # print(item)
                items.extend(item)
            # print(items) 
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy chủ tàu hợp lệ!')
        elif query_type == '4':
            users = BangThuyenTruong.objects.filter(Q(HoTen__icontains=query)).order_by('-ID').exclude(hasShip=False)
            items = []
            for user in users:
                print(user)
                item = user.bangtau
                items.append(item)
            # thuyentruong = BangThuyenTruong.objects.first()
            # bangtau_instance = thuyentruong.bangtau
            # print(bangtau_instance)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thuyền trưởng hợp lệ!')
        elif query_type == '5':
            devices = BangThietBiNhatKyKhaiThac.objects.filter(Q(SerialNumber__icontains=query)).exclude(is_active=False)
            items = []
            for d in devices:
                item = d.bangtau
                items.append(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thiết bị hợp lệ!')
        elif query_type == '6':
            items = BangTau.objects.filter(Q(IMO__icontains=query) | Q(SoDangKy__icontains=query))  
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin IMO hoặc Số đăng ký hợp lệ!')

        return render(request, 'core/device.html', {
            'titles': titles,
            'items': items, 
        }, status=200)
    elif request.user.user_type == '2':
        items = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca)).order_by('-ID')

        if query_type == '1':
            items = BangTau.objects.filter(Q(IMO__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca)).order_by('-ID')
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy IMO hợp lệ!')
        elif query_type == '2':
            items = BangTau.objects.filter(Q(SoDangKy__icontains=query) & Q(CangCaDangKy=request.user.staff.cangca)).order_by('-ID') 
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy số đăng ký tàu hợp lệ!')
        elif query_type == '3':
            users = BangChuTau.objects.filter(Q(HoTen__icontains=query)).order_by('-ID')
            # print(users)
            items = []
            for user in users:
                item = user.bangtau_chutau
                for i in item:
                    if i.CangCaDangKy.ID == request.user.staff.cangca.ID:
                        items.append(item)
            # print(items) 
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy chủ tàu hợp lệ!')
        elif query_type == '4':
            users = BangThuyenTruong.objects.filter(Q(HoTen__icontains=query)).order_by('-ID').exclude(hasShip=False)
            items = []
            for user in users:
                print(user)
                item = user.bangtau
                if item.CangCaDangKy.ID == request.user.staff.cangca.ID:
                    items.append(item)
            # thuyentruong = BangThuyenTruong.objects.first()
            # bangtau_instance = thuyentruong.bangtau
            # print(bangtau_instance)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thuyền trưởng hợp lệ!')
        elif query_type == '5':
            devices = BangThietBiNhatKyKhaiThac.objects.filter(Q(SerialNumber__icontains=query)).exclude(is_active=False)
            items = []
            for d in devices:
                item = d.bangtau
                if item.CangCaDangKy.ID == request.user.staff.cangca.ID:
                    items.append(item)
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin thiết bị hợp lệ!')
        elif query_type == '6':
            items = BangTau.objects.filter(Q(IMO__icontains=query) | Q(SoDangKy__icontains=query))  
            if len(items) == 0:
                messages.info(request, f'Không tìm thấy thông tin IMO hoặc Số đăng ký hợp lệ!')

        return render(request, 'core/device.html', {
            'titles': titles,
            'items': items, 
        }, status=200) 
    else:
        return render(request, '403.html', {}, status=403)
    

# download device data
@login_required(login_url='/login/')
def download_device_data(request, number):
    if request.user.user_type == '1' or request.user.is_staff:
        number = int(number)
        wb = Workbook()
        ws = wb.active # worksheet
        if number == '1':
            qty = 50
        elif number == '2':
            qty = 100
        elif number == '3':
            qty = 200
        elif number == '4':
            qty = 500
        else:
            qty = None 
        if qty is not None:
            ship_list = BangTau.objects.all()[:qty]
        else:
            ship_list = BangTau.objects.all()

        ws["A1"] = "STT"
        ws["B1"] = "IMO"
        ws["C1"] = "Số hiệu tàu"
        ws["D1"] = "Chủ tàu"
        ws["E1"] = "Loại tàu"
        ws["F1"] = "Nơi đăng ký"
        ws["G1"] = "Thuyền trưởng"
        ws["H1"] = "Cảng cá ĐK"
        ws["I1"] = "TBNKKT" 

        for i, ship in enumerate(ship_list, start=2):
            if ship is not None:
                ws[f"A{i}"] = i - 1
                ws[f"B{i}"] = ship.IMO
                ws[f"C{i}"] = ship.SoDangKy
                ws[f"D{i}"] = ship.IDChuTau.HoTen
                ws[f"E{i}"] = ship.LoaiTau.IDLoaiTau
                ws[f"F{i}"] = ship.NoiDangKy.TenTiengViet
                ws[f"G{i}"] = ship.IDThuyenTruong.HoTen
                ws[f"H{i}"] = ship.CangCaDangKy.Ten
                ws[f"I{i}"] = ship.IDDevice.SerialNumber
            else:
                ws[f"A{i}"] = ''
                ws[f"B{i}"] = ''
                ws[f"C{i}"] = ''
                ws[f"D{i}"] = ''
                ws[f"E{i}"] = ''
                ws[f"F{i}"] = ''
                ws[f"G{i}"] = ''
                ws[f"H{i}"] = ''
                ws[f"I{i}"] = ''

        _, filepath = tempfile.mkstemp(suffix='.xlsx')
        wb.save(filepath)

        with open(filepath, 'rb') as f:
            excel_data = f.read()
        
        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
        return response
    elif request.user.user_type == '2':
        number = int(number)
        wb = Workbook()
        ws = wb.active # worksheet
        if number == '1':
            qty = 50
        elif number == '2':
            qty = 100
        elif number == '3':
            qty = 200
        elif number == '4':
            qty = 500
        else:
            qty = None 
        if qty is not None:
            ship_list = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca))[:qty]
        else:
            ship_list = BangTau.objects.filter(Q(CangCaDangKy=request.user.staff.cangca))

        ws["A1"] = "STT"
        ws["B1"] = "IMO"
        ws["C1"] = "Số hiệu tàu"
        ws["D1"] = "Chủ tàu"
        ws["E1"] = "Loại tàu"
        ws["F1"] = "Nơi đăng ký"
        ws["G1"] = "Thuyền trưởng"
        ws["H1"] = "Cảng cá ĐK"
        ws["I1"] = "TBNKKT" 

        for i, ship in enumerate(ship_list, start=2):
            ws[f"A{i}"] = i - 1
            ws[f"B{i}"] = ship.IMO
            ws[f"C{i}"] = "********"
            ws[f"D{i}"] = ship.IDChuTau.HoTen
            ws[f"E{i}"] = ship.LoaiTau.IDLoaiTau
            ws[f"F{i}"] = ship.NoiDangKy.TenTiengViet
            ws[f"G{i}"] = ship.IDThuyenTruong.HoTen
            ws[f"H{i}"] = ship.CangCaDangKy.Ten
            ws[f"I{i}"] = ship.IDDevice.SerialNumber

        _, filepath = tempfile.mkstemp(suffix='.xlsx')
        wb.save(filepath)

        with open(filepath, 'rb') as f:
            excel_data = f.read()
        
        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
        return response 
    else:
        return render(request, '403.html', {}, status=403)


# quản lý thông tin chủ tàu, thuyền trưởng
@login_required(login_url='/login/')
def shipowners_view(request):
    if request.user.user_type == '1' or request.user.is_staff:
        titles = ["STT", "Họ tên", "Số CMND(CCCD)", "Địa chỉ", "Trạm bờ", "Email", "Số điện thoại", "Người dùng", "Thao tác"]
        captains = list(BangThuyenTruong.objects.all().order_by('HoTen'))
        shipowners = list(BangChuTau.objects.all().order_by('HoTen'))

        for captain in captains:
            captain.type = 'captain'
        for shipowner in shipowners:
            shipowner.type = 'shipowner'    
        
        combined_list = captains + shipowners
        # random.shuffle(combined_list)
        return render(request, 'core/shipowners.html', {
            'titles': titles,
            'items': combined_list,
        }, status=200)
    elif request.user.user_type == '2':
        titles = ["STT", "Họ tên", "Số CMND(CCCD)", "Địa chỉ", "Trạm bờ", "Email", "Số điện thoại", "Người dùng", "Thao tác"]
        captains = []
        shipowners = []
        ships = BangTau.objects.filter(CangCaDangKy=request.user.staff.cangca)
        if ships is not None:
            for ship in ships:
                if ship.IDChuTau:
                    shipowners.append(ship.IDChuTau)
                if ship.IDThuyenTruong:
                    captains.append(ship.IDThuyenTruong)
        for captain in captains:
            captain.type = 'captain'
        for shipowner in shipowners:
            shipowner.type = 'shipowner' 
        
        combined_list = list(set(captains)) + list(set(shipowners))
        combined_list = sorted(combined_list, key=operator.attrgetter('HoTen'))
        return render(request, 'core/shipowners.html', {
            'titles': titles,
            'items': combined_list,
        }, status=200)
    else:
        return render(request, '403.html', {}, status=403) 


# tìm kiếm thông tin chủ tàu, thuyền trưởng
@login_required(login_url='/login/')
def search_shipowners_view(request):
    query = request.GET.get('q')
    query_type = request.GET.get('query-type')
    titles = ["STT", "Họ tên", "Số CMND(CCCD)", "Địa chỉ", "Trạm bờ", "Email", "Số điện thoại", "Người dùng", "Thao tác"]
    captains = list(BangThuyenTruong.objects.all())
    shipowners = list(BangChuTau.objects.all())

    if query_type == '1':
        captains = []
        shipowners = list(BangChuTau.objects.filter(Q(HoTen__icontains=query)))
        if len(captains) == 0 and len(shipowners) == 0:
            messages.info(request, f"Không tìm thấy thông tin họ tên chủ tàu hợp lệ với query='{query}'")
    elif query_type == '2':
        captains = list(BangThuyenTruong.objects.filter(Q(HoTen__icontains=query)))
        shipowners = []
        if len(captains) == 0 and len(shipowners) == 0:
            messages.info(request, f"Không tìm thấy thông tin họ tên thuyền trưởng hợp lệ với query='{query}'")
    elif query_type == '3':
        captains = []
        shipowners = list(BangChuTau.objects.filter(Q(IDChuTau__icontains=query)))
        if len(captains) == 0 and len(shipowners) == 0:
            messages.info(request, f"Không tìm thấy thông tin mã chủ tàu hợp lệ với query='{query}'")
    elif query_type == '4':
        captains = list(BangThuyenTruong.objects.filter(Q(IDThuyenTruong__icontains=query)))
        shipowners = []
        if len(captains) == 0 and len(shipowners) == 0:
            messages.info(request, f"Không tìm thấy thông tin mã thuyền trưởng hợp lệ với query='{query}'") 
    elif query_type == '5':
        captains = list(BangThuyenTruong.objects.filter(Q(CMND_CCCD__icontains=query)))
        shipowners = list(BangChuTau.objects.filter(Q(CMND_CCCD__icontains=query)))
        if len(captains) == 0 and len(shipowners) == 0:
            messages.info(request, f"Không tìm thấy thông tin CMND/CCCD hợp lệ với query='{query}'") 
    elif query_type == '6':
        captains = list(BangThuyenTruong.objects.filter(Q(DienThoai__icontains=query)))
        shipowners = list(BangChuTau.objects.filter(Q(DienThoai__icontains=query)))
        if len(captains) == 0 and len(shipowners) == 0:
            messages.info(request, f"Không tìm thấy thông tin số điện thoại hợp lệ với query='{query}'") 
    elif query_type == '7':
        captains = list(BangThuyenTruong.objects.filter(Q(HoTen__icontains=query) | Q(IDThuyenTruong__icontains=query) | Q(CMND_CCCD__icontains=query) | Q(DienThoai__icontains=query)))
        shipowners = list(BangChuTau.objects.filter(Q(HoTen__icontains=query) | Q(IDChuTau__icontains=query) | Q(CMND_CCCD__icontains=query) | Q(DienThoai__icontains=query)))
        if len(captains) == 0 and len(shipowners) == 0:
            messages.info(request, f"Không tìm thấy thông tin truy vấn hợp lệ với query='{query}'") 


    for captain in captains:
        captain.type = 'captain'
    for shipowner in shipowners:
        shipowner.type = 'shipowner'

    combined_list = captains + shipowners
    random.shuffle(combined_list)
    return render(request, 'core/shipowners.html', {
        'titles': titles,
        'items': combined_list,
    }, status=200)


# thêm chủ tàu, thuyền trưởng
@login_required(login_url='/login/')
def add_shipowners_view(request):
    if request.method == 'POST':
        HoTen = request.POST.get('hoTen')
        CMND_CCCD = request.POST.get('CMND-CCCD')
        NgaySinh = request.POST.get('ngaySinh')
        DiaChi = request.POST.get('diaChi')
        DienThoai = request.POST.get('dienThoai')
        Fax = request.POST.get('fax')
        Email = request.POST.get('email')
        try:
            Huyen = BangDonViHanhChinhCapHuyen.objects.get(MaHuyen=request.POST['maHuyen'])
        except Exception as e:
            messages.error(request, f"Không tồn tại huyện có mã tỉnh '{request.POST['maHuyen']}'")
            return redirect('add-shipowners')
        ChucDanh = request.POST.get('chucDanh')

        if ChucDanh == '1':
            new_shipowner = BangChuTau.objects.create(
                HoTen=HoTen,
                CMND_CCCD=CMND_CCCD,
                NgaySinh=NgaySinh,
                DiaChi=DiaChi,
                DienThoai=DienThoai,
                Fax=Fax,
                Email=Email,
                MaHuyen=Huyen
            )
            messages.success(request, f"Thêm thông tin chủ tàu mới thành công!!!")
            return redirect('shipowners-view')
        elif ChucDanh == '2':
            new_captain = BangThuyenTruong.objects.create(
                HoTen=HoTen,
                CMND_CCCD=CMND_CCCD,
                NgaySinh=NgaySinh,
                DiaChi=DiaChi,
                DienThoai=DienThoai,
                Fax=Fax,
                Email=Email,
                MaHuyen=Huyen
            )
            messages.success(request, f"Thêm thông tin thuyền trưởng mới thành công!!!")
            return redirect('shipowners-view')
        else:
            messages.error(request, "Lỗi không đúng định dạng dữ liệu yêu cầu!!!")
            return redirect('shipowners-view')
    distric_list = BangDonViHanhChinhCapHuyen.objects.all()
    return render(request, 'core/add-new-shipowners.html', {
        'distric_list': distric_list,
    }, status=200) 


# sửa thông tin chủ tàu, thuyền trưởng
@login_required(login_url='/login/')
def edit_shipowners_view(request, pk, user_type):
    if request.method == 'POST':

        Huyen = BangDonViHanhChinhCapHuyen.objects.get(pk=request.POST['maHuyen'])
        HoTen = request.POST.get('hoTen')
        CMND_CCCD = request.POST.get('CMND-CCCD')
        NgaySinh = request.POST.get('ngaySinh')
        DiaChi = request.POST.get('diaChi')
        DienThoai = request.POST.get('dienThoai')
        Fax = request.POST.get('fax')
        Email = request.POST.get('email')

        if user_type == 'captain':
            try:
                captain = BangThuyenTruong.objects.get(pk=pk)
            except Exception as e:
                messages.error(request, f"Không tìm thấy thông tin thuyền trưởng với ID = {pk}")
            captain.HoTen = HoTen
            captain.CMND_CCCD = CMND_CCCD
            captain.NgaySinh = NgaySinh
            captain.DiaChi = DiaChi
            captain.DienThoai = DienThoai
            captain.Fax = Fax 
            captain.Email = Email
            captain.Huyen = Huyen
            captain.save()
            
            messages.info(request, f"Cập nhật thông tin thuyền trưởng {HoTen} - {CMND_CCCD} thành công!")
            return redirect('shipowners-view')
        elif user_type == 'shipowner':
            try:
                shipowner = BangChuTau.objects.get(pk=pk)
            except Exception as e:
                messages.error(request, f"Không tìm thấy thông tin thuyền trưởng với ID = {pk}")
            shipowner.HoTen = HoTen
            shipowner.CMND_CCCD = CMND_CCCD
            shipowner.NgaySinh = NgaySinh
            shipowner.DiaChi = DiaChi
            shipowner.DienThoai = DienThoai
            shipowner.Fax = Fax 
            shipowner.Email = Email
            shipowner.Huyen = Huyen
            shipowner.save()
            
            messages.info(request, f"Cập nhật thông tin chủ tàu {HoTen} - {CMND_CCCD} thành công!")
            return redirect('shipowners-view')
        else:
            pass # return 404 page

    distric_list = BangDonViHanhChinhCapHuyen.objects.all()
    if user_type == 'captain':
        try:
            captain = BangThuyenTruong.objects.get(pk=pk)
        except Exception as e:
            messages.error(request, "Không tìm thấy thông tin thuyền trưởng!!!")
            return redirect('shipowners-view') 

        return render(request, 'core/edit-shipowners.html', {
            'item': captain,
            'user_type': 'captain',
            'distric_list': distric_list,
        }, status=200)
    elif user_type == 'shipowner':
        try:
            shipowner = BangChuTau.objects.get(pk=pk)
        except Exception as e:
            messages.error(request, "Không tìm thấy thông tin chủ tàu!!!")
            return redirect('shipowners-view') 
        
        return render(request, 'core/edit-shipowners.html', {
            'item': shipowner,
            'user_type': 'shipowner',
            'distric_list': distric_list,
        }, status=200)
    else:
        messages.error(request, "Lỗi truy vấn!!!")
        return redirect('shipowners-view') 


# xóa thông tin tàu hoặc thuyền trưởng
@login_required(login_url='/login/')
def delete_shipowners_view(request, pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        # print(data)
        if data['userType'] == 'captain':
            try:
                captain = BangThuyenTruong.objects.get(pk=pk)
            except BangThuyenTruong.DoesNotExist:
                return JsonResponse({
                    'message': f"Không tìm thấy thông tin thuyền trưởng với id = '{pk}'",
                    'success': False
                }, status=404)
            ship = BangTau.objects.filter(IDThuyenTruong=captain)
            if ship.exists():
                return JsonResponse({
                    'message': 'Không được phép xóa thuyền trưởng khi chưa xóa tàu liên kết!!!',
                    'success': False
                })
            else:
                captain.delete()
                return JsonResponse({
                    'message': "Xóa thông tin thuyền trưởng thành công!!!",
                    'success': True
                })
        elif data['userType'] == 'shipowner':
            try:
                shipowner = BangChuTau.objects.get(pk=pk)
            except BangChuTau.DoesNotExist:
                return JsonResponse({
                    'message': f"Không tìm thấy thông tin chủ tàu với id = '{pk}'",
                    'success': False
                }, status=404)

            ship = BangTau.objects.filter(IDChuTau=shipowner)
            if ship.exists():
                return JsonResponse({
                    'message': 'Không được phép xóa chủ tàu khi chưa xóa tàu liên kết!!!',
                    'success': False
                })
            else:
                shipowner.delete()
                return JsonResponse({
                    'message': "Xóa thông tin chủ tàu thành công!!!",
                    'success': True
                })
        else:
            return JsonResponse({
                'message': 'User Type không hợp lệ',
                'success': False
            })
    else:
        return JsonResponse({
            'message': 'Method not allowed!!',
            'success': False
        }, status=405)
    

# quản lý tài khoản
@login_required(login_url='/login/')
def account_view(request):
    if request.user.user_type == '1' or request.user.is_staff:
        titles = ["STT", "Tên tài khoản", "Cấp độ tài khoản", "Tình trạng", "Loại tài khoản", "Địa chỉ", "Email", "Số điện thoại", "Thao tác"]
        UserModel = get_user_model()
        items = UserModel.objects.all().order_by('username').exclude(username='admin')
        return render(request, 'core/account.html', {
            'titles': titles,
            'items': items,
        }, status=200)
    else:
        return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def add_new_account_view(request):
    if request.user.user_type == '1' or request.user.is_staff:
        if request.method == 'POST':
            UserModel = get_user_model()
            first_name = request.POST.get('firstName')
            last_name = request.POST.get('lastName')

            username = request.POST.get('username')
            # check username xem đã tồn tại hay chưa?
            if UserModel.objects.filter(username=username).exists():
                messages.error(request, f"Tên người dùng '{username}' đã tồn tại")
                return redirect('add-account')
            email = request.POST.get('email')
            # check email đã tồn tại hay chưa?
            if UserModel.objects.filter(email=email).exists():
                messages.error(request, f"Email '{email}' đã tồn tại")
                return redirect('add-account')

            password = request.POST.get('password')
            userType = request.POST.get('levelManager')
            # print(userType)
            hash_password = make_password(password)
            user = CustomUser.objects.create(
                username=username, 
                password=hash_password,
                email=email,
                last_name=last_name,
                first_name=first_name,
                user_type=userType
            )

            # print(user.user_type, type(user.user_type))
            # Admin thêm is_staff bằng True
            if user.user_type == '1':
                # print("TEST 1")
                user.is_staff = True
                user.save() 

            # Quản lý cảng thêm phần cảng
            if user.user_type == '2':
                # print("TEST")
                pk = request.POST.get('tenCangCa')
                print(pk)
                try:
                    gateName = BangCangCa.objects.get(pk=pk)
                except BangCangCa.DoesNotExist:
                    messages.error(request, f"Lỗi không tìm thấy thông tin")
                    return redirect('account-view')
                
                staff = Staff.objects.create(admin=user)
                staff.cangca = gateName
                staff.save()
                user.save()
                # print(user.staff.cangca)


            # Quản lý thiết bị

            messages.success(request, "Tạo user thành công!")
            return redirect('account-view')     
        else:
            gate_list = BangCangCa.objects.all().order_by('Ten')
            return render(request, 'core/add-new-account.html', {'gate_list': gate_list}, status=200) 
    else:
        return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def delete_account_view(request, pk):
    if request.user.user_type == '1' or request.user.is_staff:
        try:
            UserModel = get_user_model()
            user = UserModel.objects.get(pk=pk)
            user.delete()

            messages.success(request, f"Xóa user thành công!!")
            return redirect('account-view')
        except UserModel.DoesNotExist:
            messages.error(request, f"User với id = '{pk}' không tồn tại!!")
            return redirect('account-view')
    else:
        return render(request, '403.html', {}, status=403) 


@login_required(login_url='/login/')
def edit_account_view(request, pk):
    if request.user.user_type == '1' or request.user.is_staff:
        if request.method == 'POST':
            UserModel = get_user_model()
            username = request.POST.get('username')
            first_name = request.POST.get('firstName')
            last_name = request.POST.get('lastName')
            password = request.POST.get('password')
            email = request.POST.get('email')

            # level_manager = request.POST.get('levelManager') 
            try:
                user = UserModel.objects.get(pk=pk)
                user.username = username
                user.first_name = first_name 
                user.last_name = last_name 
                user.email = email 
                if password:
                    user.set_password(password)
                user.save()
                messages.success(request, 'Cập nhật thông tin người dùng thành công!')
                return redirect('account-view')
            except UserModel.DoesNotExist:
                messages.error(request, "Người dùng không tồn tại")
                return redirect('account-view')        
        try:
            user = get_user_model().objects.get(pk=pk)
        except Exception as e:
            messages.error(request, f"Không tìm thấy thông tin user với id = '{pk}'")
            return redirect('index')
        gate_list = BangCangCa.objects.all()

        return render(request, 'core/edit-account.html', {'gate_list': gate_list, 'user': user}, status=200)
    else:
        return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def search_account_view(request):
    if request.user.user_type == '1' or request.user.is_staff:
        query = request.GET.get('q')
        query_type = request.GET.get('query-type') 
        titles = ["STT", "Tên tài khoản", "Cấp độ tài khoản", "Tình trạng", "Loại tài khoản", "Địa chỉ", "Email", "Số điện thoại", "Thao tác"]
        
        UserModel = get_user_model()
        fields = UserModel._meta.fields
        for field in fields:
            print(field.name)
        
        if query_type == '1':
            # nhà cung cấp thiết bị: lv3
            items = UserModel.objects.filter((Q(user_type = '3') & Q(username__icontains=query)) | (Q(user_type = '3') & Q(first_name__icontains=query)) | (Q(user_type = '3') & Q(last_name__icontains=query)) | (Q(user_type = '3') & Q(email__icontains=query)))
            if len(items) == 0:
                messages.info(request, f"Không tìm thấy user level 3 với query = '{query}'")
        elif query_type == '2':
            # quản lý cảng: lv2
            items = UserModel.objects.filter((Q(user_type = '2') & Q(username__icontains=query)) | (Q(user_type = '3') & Q(first_name__icontains=query)) | (Q(user_type = '3') & Q(last_name__icontains=query)) | (Q(user_type = '3') & Q(email__icontains=query)))
            if len(items) == 0:
                messages.info(request, f"Không tìm thấy user level 2 với query = '{query}'") 
        elif query_type == '3':
            # chi cục thủy sản: lv1
            items = UserModel.objects.filter((Q(user_type = '1') & Q(username__icontains=query)) | (Q(user_type = '3') & Q(first_name__icontains=query)) | (Q(user_type = '3') & Q(last_name__icontains=query)) | (Q(user_type = '3') & Q(email__icontains=query)))
            if len(items) == 0:
                messages.info(request, f"Không tìm thấy user level 1 với query = '{query}'") 
        else:
            items = UserModel.objects.all()

        return render(request, 'core/account.html', {
            'titles': titles,
            'items': items,
        }, status=200)
    else:
        return render(request, '403.html', {}, status=403)
 

@login_required(login_url='/login/')
def add_new_equipment_view(request):
    if request.user.user_type == '3':
        if request.method == 'POST':
            SerialNumber = request.POST.get('serialNumber')
            NgaySanXuat = request.POST.get('ngaySanXuat')
            FWVersion = request.POST.get('FWVersion')

            equipment = BangThietBiNhatKyKhaiThac.objects.create(
                SerialNumber=SerialNumber,
                NgaySanXuat=NgaySanXuat,
                FWVersion=FWVersion
            )

            messages.info(request, "Tạo thiết bị nhật ký khai thác thành công!!!")
            return redirect('index')
        else:
            return render(request, 'core/add-equipment.html', {}, status=200)
    else:
        return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def edit_equipment_view(request, pk):
    if request.user.user_type == '3':
        if request.method == 'POST':
            SerialNumber = request.POST.get('serialNumber')
            NgaySanXuat = request.POST.get('ngaySanXuat')
            FWVersion = request.POST.get('FWVersion')

            try:
                equipment = BangThietBiNhatKyKhaiThac.objects.get(pk=pk)
            except BangThietBiNhatKyKhaiThac.DoesNotExist:
                messages.error(request, f"Không tìm thấy thiết bị với ID = '{pk}'")
                return redirect('index')
            
            equipment.SerialNumber = SerialNumber
            equipment.NgaySanXuat = NgaySanXuat 
            equipment.FWVersion = FWVersion 
            equipment.save()

            messages.info(request, f"Cập nhật thông tin thiết bị thành công!")
            return redirect('index')
        else:
            try:
                equipment = BangThietBiNhatKyKhaiThac.objects.get(pk=pk)
            except BangThietBiNhatKyKhaiThac.DoesNotExist:
                messages.error(request, f"Không tìm thấy thiết bị với ID = '{pk}'")
                return redirect('index')
            return render(request, 'core/edit-equipment.html', {'equipment': equipment}, status=200)
    else:
        return render(request, '403.html', {}, status=403)


@login_required(login_url='/login/')
def delete_equipment_view(request, pk):
    if request.user.user_type == '3':
        try:
            equipment = BangThietBiNhatKyKhaiThac.objects.get(pk=pk)
        except BangThietBiNhatKyKhaiThac.DoesNotExist:
            messages.error(request, f"Không tìm thấy thiết bị với ID = '{pk}'")
            return redirect('index')
        equipment.delete()
        messages.info(request, f"Xóa thiết bị với ID = {pk} thành công!!")
        return redirect('index')
    else:
        return render(request, '403.html', {}, status=403)


# @login_required(login_url='/login/')
# def check_valid_equipment_api(request):
#     if request.user.user_type == '3':
#         if request.method == 'POST':
#             data = json.load(request.body)
#             serial_number = data['SerialNumber']
#             equipment = BangThietBiNhatKyKhaiThac.objects.filter(SerialNumber=serial_number)
#             if equipment is not None:
#                 return JsonResponse({
#                     'message': 'Serial Number already exist!!!',
#                     'success': False
#                 }) 
#             else:
#                 return JsonResponse({
#                     'message': 'Validate form successfully!!!',
#                     'success': True
#                 })
#         else:
#             return JsonResponse({
#                 'message': 'Method not allowed',
#                 'success': False
#             }, status=405)
#     else:
#         return JsonResponse({
#             'message': "Error: User permission",
#             'success': False,
#         }, status=403)


@login_required(login_url='/login/')
def search_equipment_view(request):
    if request.user.user_type == '3':
        query = request.GET.get('q')
        query_type = request.GET.get('query-type')
        titles = ["STT", "Serial Number", "Ngày sản xuất", "Version", "Mã tàu", "Trạng thái", "Thao tác"]
        equipments = BangThietBiNhatKyKhaiThac.objects.all()

        # mã thiết bị
        if query_type == '1':
            equipments = BangThietBiNhatKyKhaiThac.objects.filter(Q(IDThietBi__icontains=query)) 
            if len(list(equipments)) == 0:
                messages.info(request, f"Không tìm thấy thông tin thiết bị liên quan đến Mã thiết bị với query='{query}'")
        
        # serial number 
        elif query_type == '2':
            equipments = BangThietBiNhatKyKhaiThac.objects.filter(Q(SerialNumber__icontains=query))
            if len(list(equipments)) == 0:
                messages.info(request, f"Không tìm thấy thông tin thiết bị liên quan đến Serial Number với query='{query}'")
            
        # mã tàu
        # elif query_type == '3':
        #     ships = BangTau.objects.filter(Q(SoDangKy__icontains=query))
        #     equipments = []
        #     for ship in ships:
        #         # print(ship.IDDevice)
        #         item = BangThietBiNhatKyKhaiThac.objects.filter(pk=ship.IDDevice.ID)
        #         equipments.append(item)
        #     if len(list(equipments)) == 0:
        #         messages.info(request, f"Không tìm thấy thông tin thiết bị liên quan đến mã tàu với query='{query}'")
        
        # print(equipments)
        return render(request, 'core/index.html', {
            'titles': titles,
            'items': equipments
        }, status=200)
    else:
        return render(request, '403.html', {}, status=403)


@csrf_exempt
def login_mobile(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')

            if username == 'admin':
                if password == '@Long1705':
                    return JsonResponse({
                        'message': 'Login successfully!', 
                        'passed': True, 
                        'username': 'Long02'
                    }, status=200)
                else:
                    return JsonResponse({
                        'message': 'Password is not correct!',
                        'passed': False,
                        'username': 'Anonymous'
                    }, staus=401) 
            else:
                return JsonResponse({
                    'message': 'Username does not exits!',
                    'passed': False,
                    'username': 'Anonymous'
                })
        except Exception as e:
            return JsonResponse({
                'message': 'Server error!',
                'passed': False,
                'username': 'Anonymous'
            })
    else:
        return JsonResponse({
            'message': 'Method not allowed!',
            'passed': False,
            'username': 'Anonymous'
        })

    
def test_api_mobile(request):
    return JsonResponse({
        'message': 'Hello World!',
        'status': 200
    }, status=200)


def ship_name_list(request):
    return JsonResponse({
        'message': 'Danh sách tàu cá',
        'status': 200,
        'data': [
            {
                'so_hieu_tau': 'KH-12235',
                'chu_tau': 'Nguyễn Văn Long'
            },
            {
                'so_hieu_tau': 'BĐ-12935',
                'chu_tau': 'Nguyễn Văn Long'
            },
            {
                'so_hieu_tau': 'HP-99999',
                'chu_tau': 'Nguyễn Văn Long'
            },
            {
                'so_hieu_tau': 'HN-29299',
                'chu_tau': 'Nguyễn Văn Long'
            },
        ]
    }, status=200)


def ship_realtime_location(request):
    now = datetime.now()
    formatted_date = now.strftime("%d %b, %Y %H:%M:%S")
    print(formatted_date)
    # 16.344088, 108.599521 [16.408051, 108.489041, 16.415336, 108.251537, 16.583879, 107.982372]
    # 13.811880, 109.594370 [13.658198, 109.478250, 13.509807, 109.674731, 13.100527, 109.479232]
    # 12.963734, 109.654796 [12.824108, 109.441995, 12.655258, 109.672265, 12.336105, 109.328733]
    # 11.457285, 109.207992 [11.636259, 109.196456, 11.819430, 109.472452, 11.992313, 109.358127]
    return JsonResponse({
        'message': 'Vị trí tàu cá hiện tại.',
        'status': 200,
        'ship_list': [
            {
                'so_hieu_tau': 'KH-12235',
                'thuyen_truong': 'Nguyễn Văn A',
                'chu_tau': 'Nguyễn Văn Long',
                'loai_thiet_bi': 'VSM',
                'ten_thiet_bi': 'VHK-S',
                'IMO': '1705',
                'ngay_dang_ky': '2017-01-01',
                'ngay_het_han_dang_ky': '2027-01-01',
                'so_kep_chi': 'VHK-123455',
                'ngay_niem_phong': '20/02/2006',
                'location': {
                    'lat': 16.344088,
                    'lng': 108.599521,
                },
                'time': formatted_date,
            },
            {
                'so_hieu_tau': 'BĐ-12935',
                'thuyen_truong': 'Nguyễn Văn A',
                'chu_tau': 'Nguyễn Văn Long',
                'loai_thiet_bi': 'VSM',
                'ten_thiet_bi': 'VHK-S',
                'IMO': '1777',
                'ngay_dang_ky': '2017-01-01',
                'ngay_het_han_dang_ky': '2027-01-01',
                'so_kep_chi': 'VHK-129455',
                'ngay_niem_phong': '20/02/2006',
                'location': {
                    'lat': 13.811880,
                    'lng': 109.594370,
                },
                'time': formatted_date,
            },
            {
                'so_hieu_tau': 'HP-99999',
                'thuyen_truong': 'Nguyễn Văn A',
                'chu_tau': 'Nguyễn Văn Long',
                'loai_thiet_bi': 'VSM',
                'ten_thiet_bi': 'VHK-S',
                'IMO': '2000',
                'ngay_dang_ky': '2017-01-01',
                'ngay_het_han_dang_ky': '2027-01-01',
                'so_kep_chi': 'VHK-129955',
                'ngay_niem_phong': '20/02/2006',
                'location': {
                    'lat': 12.963734,
                    'lng': 109.654796,
                },
                'time': formatted_date,
            },
            {
                'so_hieu_tau': 'HN-29299',
                'thuyen_truong': 'Nguyễn Văn A',
                'chu_tau': 'Nguyễn Văn Long',
                'loai_thiet_bi': 'VSM',
                'ten_thiet_bi': 'VHK-S',
                'IMO': '1723',
                'ngay_dang_ky': '2017-01-01',
                'ngay_het_han_dang_ky': '2027-01-01',
                'so_kep_chi': 'VHK-123685',
                'ngay_niem_phong': '20/02/2006',
                'location': {
                    'lat': 11.457285,
                    'lng': 109.207992,
                },
                'time': formatted_date,
            },
        ],
    }, status=200) 


def ship_location_log(request):
    return JsonResponse({
        'message': 'Lịch sử vị trí tàu cá',
        'status': 200,
        'log': {
            'tc1': {
                'thong_tin_chung': {
                    'so_hieu_tau': 'KH-12235',
                    'thuyen_truong': 'Nguyễn Văn A',
                    'chu_tau': 'Nguyễn Văn Long',
                    'loai_thiet_bi': 'VSM',
                    'ten_thiet_bi': 'VHK-S',
                    'IMO': '1705',
                    'ngay_dang_ky': '2017-01-01',
                    'ngay_het_han_dang_ky': '2027-01-01',
                    'so_kep_chi': 'VHK-123455',
                    'ngay_niem_phong': '20/02/2006',
                },
                'journal': [
                    {
                        'lat': 16.408051,
                        'lng': 108.489041,
                        'date': '13-01-2024',
                    },
                    {
                        'lat': 16.415336,
                        'lng': 108.251537,
                        'date': '29-12-2023',
                    },
                    {
                        'lat': 16.583879,
                        'lng': 107.982372,
                        'date': '29-11-2023',
                    },
                ]
            },


            'tc2': {
                'thong_tin_chung': {
                    'so_hieu_tau': 'BĐ-12935',
                    'thuyen_truong': 'Nguyễn Văn A',
                    'chu_tau': 'Nguyễn Văn Long',
                    'loai_thiet_bi': 'VSM',
                    'ten_thiet_bi': 'VHK-S',
                    'IMO': '1777',
                    'ngay_dang_ky': '2017-01-01',
                    'ngay_het_han_dang_ky': '2027-01-01',
                    'so_kep_chi': 'VHK-129455',
                    'ngay_niem_phong': '20/02/2006',
                },
                'journal': [
                    {
                        'lat': 13.658198,
                        'lng': 109.478250,
                        'date': '13-01-2024',
                    },
                    {
                        'lat': 13.509807,
                        'lng': 109.674731,
                        'date': '29-12-2023',
                    },
                    {
                        'lat': 13.100527,
                        'lng': 109.479232,
                        'date': '29-11-2023',
                    },
                ]
            },


            'tc3': {
                'thong_tin_chung': {
                    'so_hieu_tau': 'HP-99999',
                    'thuyen_truong': 'Nguyễn Văn A',
                    'chu_tau': 'Nguyễn Văn Long',
                    'loai_thiet_bi': 'VSM',
                    'ten_thiet_bi': 'VHK-S',
                    'IMO': '2000',
                    'ngay_dang_ky': '2017-01-01',
                    'ngay_het_han_dang_ky': '2027-01-01',
                    'so_kep_chi': 'VHK-129955',
                    'ngay_niem_phong': '20/02/2006',
                },
                'journal': [
                    {
                        'lat': 12.824108,
                        'lng': 109.441995,
                        'date': '13-01-2024',
                    },
                    {
                        'lat': 12.655258,
                        'lng': 109.672265,
                        'date': '29-12-2023',
                    },
                    {
                        'lat': 12.336105,
                        'lng': 109.328733,
                        'date': '29-11-2023',
                    },
                ]
            },


            'tc4': {
                'thong_tin_chung': {
                    'so_hieu_tau': 'HN-29299',
                    'thuyen_truong': 'Nguyễn Văn A',
                    'chu_tau': 'Nguyễn Văn Long',
                    'loai_thiet_bi': 'VSM',
                    'ten_thiet_bi': 'VHK-S',
                    'IMO': '1723',
                    'ngay_dang_ky': '2017-01-01',
                    'ngay_het_han_dang_ky': '2027-01-01',
                    'so_kep_chi': 'VHK-123685',
                    'ngay_niem_phong': '20/02/2006',
                },
                'journal': [
                    {
                        'lat': 11.636259,
                        'lng': 109.196456,
                        'date': '13-01-2024',
                    },
                    {
                        'lat': 11.819430,
                        'lng': 109.472452,
                        'date': '29-12-2023',
                    },
                    {
                        'lat': 11.992313,
                        'lng': 109.358127,
                        'date': '29-11-2023',
                    },
                ]
            },
        },  
    }, status=200)


def mining_log(request):
    return JsonResponse({
        'message': 'Nhật ký khai thác điện tử',
        'status': 200,
        'data': [
            {
                'so_hieu_tau': 'KH-12235',
                'danh_sach_chuyen_bien': [
                    {
                        'chuyen_bien_so': 1,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2310,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2550,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1025,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5010,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 2,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1100,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1400,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5505,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2220,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1110,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 3300,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3490,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1235,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2305,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 3,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1255,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 3555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1055,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2015,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1545,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3455,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1035,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1450,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 4,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2150,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1050,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1250,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2350,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2030,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1540,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3550,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2110,
                                    },
                                ]
                            },

                        ]
                    },
                ]
            },
            {
                'so_hieu_tau': 'HN-29299',
                'danh_sach_chuyen_bien': [
                    {
                        'chuyen_bien_so': 1,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá bơn',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá bơn',
                                        'khoi_luong': 2310,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2550,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1025,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5010,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 2,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá bò hoa vàng',
                                        'khoi_luong': 3200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1100,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1400,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá nục thuôn',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5505,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2220,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1110,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 3300,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3490,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1235,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2305,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 3,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1255,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 3555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1055,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2015,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1545,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3455,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1035,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1450,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 4,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2150,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1050,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1250,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2350,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2030,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1540,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3550,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2110,
                                    },
                                ]
                            },

                        ]
                    },
                ]
            },
            {
                'so_hieu_tau': 'HP-99999',
                'danh_sach_chuyen_bien': [
                    {
                        'chuyen_bien_so': 1,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá Hồng Lang',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Nóc Chóp',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá Nóc Chóp',
                                        'khoi_luong': 1200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2310,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2550,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Nóc Chóp',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Nóc Chóp',
                                        'khoi_luong': 1025,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5010,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 2,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1100,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1400,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5505,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2220,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1110,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 3300,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3490,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1235,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2305,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 3,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1255,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 3555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1055,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2015,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1545,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3455,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1035,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1450,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 4,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2150,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1050,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1250,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2350,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2030,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1540,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3550,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2110,
                                    },
                                ]
                            },

                        ]
                    },
                ]
            },
            {
                'so_hieu_tau': 'BĐ-12935',
                'danh_sach_chuyen_bien': [
                    {
                        'chuyen_bien_so': 1,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá Ngừ Ồ',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Ngừ Chấm',
                                        'khoi_luong': 5500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2310,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2550,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Ngừ Vằn',
                                        'khoi_luong': 2000,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1025,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Ngừ Vằn',
                                        'khoi_luong': 5010,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 2,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3200,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Ngừ Chấm',
                                        'khoi_luong': 1100,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá Ngừ Chấm',
                                        'khoi_luong': 1400,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 5505,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2220,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1110,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá Ngừ Chù',
                                        'khoi_luong': 3300,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3490,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1235,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2305,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 3,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2300,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1450,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1255,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2315,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 3555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1055,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2015,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1545,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3455,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1035,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1450,
                                    },
                                ]
                            },

                        ]
                    },
                    {
                        'chuyen_bien_so': 4,
                        'danh_sach_me_ca': [
                            {
                                'me_ca': 1,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 2150,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1050,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1500,
                                    },
                                ]
                            },
                            {
                                'me_ca': 2,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1250,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2350,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2555,
                                    },
                                ]
                            },
                            {
                                'me_ca': 3,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 2030,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 1540,
                                    },
                                ]
                            },
                            {
                                'me_ca': 4,
                                'du_lieu_me_ca': [
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây vàng',
                                        'khoi_luong': 3550,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá ngừ vây xanh',
                                        'khoi_luong': 1020,
                                    },
                                    {
                                        'ten_loai_ca': 'Cá thu',
                                        'khoi_luong': 2110,
                                    },
                                ]
                            },

                        ]
                    },
                ]
            },
        ]
    }, status=200)
    


# RESTful API
@login_required(login_url='/login/')
def get_location_view_api(request, pk):
    try:
        ship = BangTau.objects.get(pk=pk)
    except Exception as e:
        return JsonResponse({
            'status': 404,
            'message': 'Ship not Found!',
            'bundle': {}
        }, status=404)

    location = BangViTriTau.objects.filter(IDTau=ship).order_by('-Ngay')[0]
    return JsonResponse({
        'status': 200,
        'message': f'Get info of ship: {ship.SoDangKy} successfully!',
        'bundle': {
            'shipowner': ship.IDChuTau.HoTen,
            'captain': ship.IDThuyenTruong.HoTen,
            'lat': location.ViDo,
            'lng': location.KinhDo,
            'date': location.Ngay,
        }
    }, status=200)
